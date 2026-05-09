import os
import re
import warnings
from collections import OrderedDict

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")


def parse_sku_table(file_path: str) -> dict:
    """
    解析单个月份SKU表，返回结构化数据
    {
        "SKC货号-类目名": {
            "skc_name": "ATYK",
            "category": "挂毯",
            "skus": {
                "SKU_ID_1": {"总销量": 77, "结算收入金额": 218.63, ...},
                "SKU_ID_2": {...}
            },
            "skc_totals": {"总销量": 77, ...}
        }
    }
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    result = OrderedDict()
    current_skc_key = None
    current_skc_name = None
    current_category = None
    current_skus = {}
    current_skc_totals = {}
    sku_header_row = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
        cell_a = row[0].value if len(row) > 0 else None

        if cell_a is None:
            continue

        cell_a_str = str(cell_a).strip()

        if cell_a_str.startswith("SKC货号-"):
            if current_skc_key is not None:
                result[current_skc_key] = {
                    "skc_name": current_skc_name,
                    "category": current_category,
                    "skus": current_skus,
                    "skc_totals": current_skc_totals
                }

            parts = cell_a_str.replace("SKC货号-", "", 1).split("-", 1)
            current_skc_name = parts[0] if len(parts) > 0 else "未知SKC"
            current_category = parts[1] if len(parts) > 1 else "未知类目"
            current_skc_key = f"{current_skc_name}-{current_category}"
            current_skus = {}
            current_skc_totals = {}
            sku_header_row = None
            continue

        if cell_a_str == "指标" and current_skc_key is not None:
            sku_header_row = row
            continue

        if sku_header_row is not None and current_skc_key is not None:
            metric_name = cell_a_str
            if metric_name == "指标":
                continue

            total_val = row[1].value if len(row) > 1 else None
            current_skc_totals[metric_name] = total_val

            for col_idx in range(2, len(row)):
                sku_name = sku_header_row[col_idx].value
                if sku_name is None:
                    continue
                sku_name = str(sku_name).strip()
                if sku_name in ("指标", "总计"):
                    continue

                if sku_name not in current_skus:
                    current_skus[sku_name] = {}
                current_skus[sku_name][metric_name] = row[col_idx].value

    if current_skc_key is not None:
        result[current_skc_key] = {
            "skc_name": current_skc_name,
            "category": current_category,
            "skus": current_skus,
            "skc_totals": current_skc_totals
        }

    wb.close()
    return result


def merge_numeric(val1, val2):
    """合并两个数值"""
    def to_num(v):
        if v is None or v == "" or v == "None":
            return 0
        if isinstance(v, (int, float)):
            return v
        try:
            return float(str(v).replace("%", "").replace(",", ""))
        except (ValueError, TypeError):
            return 0

    return to_num(val1) + to_num(val2)


def generate_sku_summary(shop_abbr: str, months_list: list = None):
    """
    读取 配置文件_财务汇总/{shop_abbr} 下所有月份SKU表，
    按SKU汇总所有月份数据，生成汇总表到 配置文件_财务汇总/{shop_abbr}/{shop_abbr}_SKU汇总.xlsx

    表头结构：指标 | 总计 | 月份1 | 月份2 | ... | SKU1 | SKU2 | ...
    每个SKC一个区块，SKC标题行带类目名
    
    Args:
        shop_abbr: 店铺缩写
        months_list: 要融合的月份列表，如 ["2025.02", "2025.03"]。如果为None，则读取目录下所有月份SKU表
    """
    logger.info(f"开始生成店铺{shop_abbr}的SKU表汇总")
    base_dir = f"配置文件_财务汇总/{shop_abbr}"
    if not os.path.exists(base_dir):
        logger.error(f"❌ 目录不存在：{base_dir}")
        return

    all_sku_files = sorted([f for f in os.listdir(base_dir) if f.endswith("_SKU.xlsx") and not f.startswith("~$")])
    
    if months_list:
        sku_files = []
        for month in months_list:
            filename = f"{month}_SKU.xlsx"
            if filename in all_sku_files:
                sku_files.append(filename)
            else:
                logger.warning(f"⚠️ 未找到月份 {month} 的SKU表文件：{filename}")
    else:
        sku_files = all_sku_files
    
    if not sku_files:
        logger.error(f"❌ 未找到SKU表文件：{base_dir}")
        return

    all_months_data = OrderedDict()
    for sku_file in sku_files:
        month = sku_file.replace("_SKU.xlsx", "")
        file_path = os.path.join(base_dir, sku_file)
        try:
            month_data = parse_sku_table(file_path)
            all_months_data[month] = month_data
            logger.info(f"✅ 读取 {sku_file} 成功（{len(month_data)}个SKC）")
        except Exception as e:
            logger.error(f"❌ 读取 {sku_file} 失败：{e}")

    metric_order = [
        "总销量", "JIT销量", "备货销量", "结算收入金额", "平台补贴金额",
        "成本金额", "退货金额", "履约售后问题金额",
        "结算总表利润", "每单利润", "利润率", "退货金额占比",
        "售后问题金额占比", "售后履约问题率总计"
    ]

    rate_metrics = {"利润率", "退货金额占比", "售后问题金额占比", "售后履约问题率总计", "每单利润"}

    all_months = list(all_months_data.keys())

    all_skus_set = OrderedDict()
    for month, month_data in all_months_data.items():
        for skc_key, skc_info in month_data.items():
            for sku_name in skc_info["skus"]:
                if sku_name not in all_skus_set:
                    all_skus_set[sku_name] = skc_key

    output_path = os.path.join(base_dir, f"{shop_abbr}_SKU汇总.xlsx")

    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    skc_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    month_header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    normal_font = Font(size=12)

    COL_WIDTH_INDICATOR = 35
    COL_WIDTH_OTHER = 20
    COL_WIDTH_SKU = 19
    ROW_HEIGHT_TITLE = 40
    ROW_HEIGHT_HEADER = 25
    ROW_HEIGHT_DATA = 32
    ROW_HEIGHT_SPACE = 10

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        wb = writer.book
        ws = wb.create_sheet(title="SKU汇总")

        current_row = 1

        months_str = " ".join(all_months)
        ws.row_dimensions[current_row].height = ROW_HEIGHT_TITLE
        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws[f"A{current_row}"] = f"{shop_abbr} - SKU汇总 - [{months_str}]"
        ws[f"A{current_row}"].fill = title_fill
        ws[f"A{current_row}"].font = title_font
        ws[f"A{current_row}"].alignment = center_align
        current_row += 2

        all_skc_keys_ordered = OrderedDict()
        for month, month_data in all_months_data.items():
            for skc_key in month_data:
                if skc_key not in all_skc_keys_ordered:
                    all_skc_keys_ordered[skc_key] = month_data[skc_key]

        for skc_key, skc_info in all_skc_keys_ordered.items():
            skc_name = skc_info["skc_name"]
            category = skc_info["category"]

            skc_all_skus = OrderedDict()
            for month in all_months:
                month_data = all_months_data.get(month, {})
                if skc_key in month_data:
                    for sku_name in month_data[skc_key]["skus"]:
                        if sku_name not in skc_all_skus:
                            skc_all_skus[sku_name] = True

            skc_sku_list = list(skc_all_skus.keys())
            if not skc_sku_list:
                continue

            col_num = 1 + len(skc_sku_list)

            ws.row_dimensions[current_row].height = ROW_HEIGHT_TITLE
            ws.merge_cells(f"A{current_row}:F{current_row}")
            ws[f"A{current_row}"] = f"SKC货号-{skc_name}-{category}"
            ws[f"A{current_row}"].fill = title_fill
            ws[f"A{current_row}"].font = title_font
            ws[f"A{current_row}"].alignment = center_align
            current_row += 2

            ws.row_dimensions[current_row].height = ROW_HEIGHT_HEADER
            headers = ["指标"] + skc_sku_list
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.fill = skc_header_fill
                cell.border = thin_border
                cell.alignment = center_align
                cell.font = normal_font
            current_row += 1

            for metric in metric_order:
                ws.row_dimensions[current_row].height = ROW_HEIGHT_DATA
                ws.cell(row=current_row, column=1, value=metric).alignment = center_align

                for s_idx, sku_name in enumerate(skc_sku_list):
                    col = 2 + s_idx
                    val = None
                    for month in all_months:
                        month_data = all_months_data.get(month, {})
                        if skc_key in month_data and sku_name in month_data[skc_key]["skus"]:
                            sv = month_data[skc_key]["skus"][sku_name].get(metric)
                            val = merge_numeric(val, sv)
                    ws.cell(row=current_row, column=col, value=_format_cell_value(val, metric))

                for col_idx in range(1, col_num + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = thin_border
                    cell.font = normal_font
                    cell.alignment = center_align
                current_row += 1

            ws.row_dimensions[current_row].height = ROW_HEIGHT_SPACE
            current_row += 2

            ws.column_dimensions[get_column_letter(1)].width = COL_WIDTH_INDICATOR
            for col_idx in range(2, col_num + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTH_SKU

    abs_path = os.path.abspath(output_path)
    logger.info(f"\n✅ SKU汇总表生成成功！")
    logger.info(f"📁 文件路径：{abs_path}")
    logger.info(f"📊 汇总月份数：{len(all_months)} | SKC组数：{len(all_skc_keys_ordered)} | SKU总数：{len(all_skus_set)}")
    return abs_path


def _format_cell_value(value, metric: str):
    """格式化单元格值"""
    if value is None or value == "" or value == "None":
        return 0
    if isinstance(value, str):
        if "%" in value:
            return value
        try:
            value = float(value)
        except (ValueError, TypeError):
            return 0
    try:
        val = float(value)
    except (ValueError, TypeError):
        return 0

    if "销量" in metric:
        return int(round(val))
    elif any(kw in metric for kw in ["率", "占比"]):
        # merge_numeric已去掉%，原始数据已是百分比数值，直接格式化添加%符号
        return f"{val:.2f}%"
    else:
        return round(val, 2)


if __name__ == "__main__":
    import sys
    
    # 支持命令行参数：python caiwu_sku_excel.py AE
    if len(sys.argv) > 1:
        shop_abbr = sys.argv[1]
    else:
        # 默认使用 AE 店铺
        shop_abbr = "RV"
    
    logger.info(f"🚀 开始生成 {shop_abbr} 店铺的 SKU 汇总表...")
    result_path = generate_sku_summary(shop_abbr, months_list=["2026.01", "2026.02"])
    if result_path:
        logger.info(f"✅ 处理完成！")
    else:
        logger.error("❌ 处理失败！")
