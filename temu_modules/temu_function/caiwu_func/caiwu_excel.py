import os
import warnings
from typing import Dict, List, Any, Optional, Union

import numpy as np
import pandas as pd
from loguru import logger
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from temu_modules.temu_function.caiwu_func.caiwu_calculate import start_exctract_money_excel
from utils.TemuBase import get_shop_info_db

# 忽略openpyxl的默认样式警告（避免控制台冗余输出）
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")


# ====================== 通用Excel处理函数 ======================
def read_excel_data(
        file_path: str,
        sheet_name: Optional[Union[str, int]] = 0,
) -> pd.DataFrame:
    """通用Excel读取函数：读取任意Excel文件的指定工作表，返回清理后的DataFrame"""
    abs_file_path = os.path.abspath(file_path)
    if not os.path.exists(abs_file_path):
        raise FileNotFoundError(f"文件不存在：{abs_file_path}（输入路径：{file_path}）")

    try:
        df = pd.read_excel(
            io=abs_file_path,
            sheet_name=sheet_name,
            engine="openpyxl",
        )
        df_clean = df.dropna(how="all").reset_index(drop=True)
        file_name = os.path.basename(abs_file_path)
        logger.info(f"✅ 读取成功：{file_name}")
        return df_clean

    except ValueError as e:
        if "Worksheet" in str(e):
            try:
                xl = pd.ExcelFile(abs_file_path, engine="openpyxl")
                raise ValueError(
                    f"工作表不存在：{sheet_name}（文件：{os.path.basename(abs_file_path)}）\n"
                    f"可用工作表：{xl.sheet_names}"
                ) from e
            except Exception:
                raise ValueError(f"读取工作表失败：{str(e)}") from e
        else:
            raise ValueError(f"Excel读取失败：{str(e)}") from e
    except Exception as e:
        raise Exception(f"文件处理异常：{str(e)}（路径：{abs_file_path}）") from e


def calculate_column_sum(
        df: pd.DataFrame,
        target_col: str,
        decimal_places: int = 2,
        unit: str = "无",
        show_invalid: bool = False,
        where_conditions: Optional[Union[Dict[str, Any], str]] = None
) -> float:
    """通用列求和函数（支持类SQL WHERE条件过滤）"""
    df_filtered = df.copy()
    if where_conditions is not None:
        if isinstance(where_conditions, dict):
            missing_cols = [col for col in where_conditions.keys() if col not in df_filtered.columns]
            if missing_cols:
                raise KeyError(f"WHERE条件中的列不存在：{missing_cols}\n当前列名：{list(df_filtered.columns)}")
            for col, value in where_conditions.items():
                df_filtered = df_filtered[df_filtered[col] == value]
        elif isinstance(where_conditions, str):
            try:
                mask = pd.eval(f"`{df_filtered.index.name}` in df_filtered.index and ({where_conditions})",
                               local_dict={"df_filtered": df_filtered})
                df_filtered = df_filtered[mask]
            except Exception as e:
                raise ValueError(f"WHERE条件解析失败：{where_conditions}\n错误原因：{str(e)}") from e

    if target_col not in df_filtered.columns:
        raise KeyError(f"目标列不存在：'{target_col}'\n当前列名：{list(df_filtered.columns)}")

    df_with_index = df_filtered[[target_col]].reset_index(drop=False)
    df_with_index["cleaned"] = df_with_index[target_col].astype(str).str.replace(r"[^\d.-]", "", regex=True)
    df_with_index["numeric"] = pd.to_numeric(df_with_index["cleaned"], errors="coerce")

    valid_data = df_with_index[df_with_index["numeric"].notna()]
    invalid_data = df_with_index[df_with_index["numeric"].isna()]
    valid_count = len(valid_data)
    invalid_count = len(invalid_data)

    if show_invalid and invalid_count > 0:
        logger.info(f"\n❌ 发现 {invalid_count} 行无效数据：")
        for idx, row in invalid_data.head(10).iterrows():
            filtered_row = row["index"] + 1
            logger.info(f"   行{filtered_row}：原始值='{row[target_col]}' → 清理后='{row['cleaned']}'")
        if invalid_count > 10:
            logger.info(f"   ... （剩余 {invalid_count - 10} 行无效数据未显示）")

    col_total = round(valid_data["numeric"].sum(), decimal_places)
    return col_total


def get_column_unique_values(
        df: pd.DataFrame,
        target_col: str,
        sort_values: bool = True,
        drop_na: bool = True,
        where_conditions: Optional[Union[Dict[str, Any], str]] = None
) -> List[Any]:
    """通用列去重取值函数"""
    df_filtered = df.copy()
    if where_conditions is not None:
        if isinstance(where_conditions, dict):
            missing_cols = [col for col in where_conditions.keys() if col not in df_filtered.columns]
            if missing_cols:
                raise KeyError(f"WHERE条件中的列不存在：{missing_cols}")
            for col, value in where_conditions.items():
                df_filtered = df_filtered[df_filtered[col] == value]
        elif isinstance(where_conditions, str):
            try:
                mask = pd.eval(f"`{df_filtered.index.name}` in df_filtered.index and ({where_conditions})",
                               local_dict={"df_filtered": df_filtered})
                df_filtered = df_filtered[mask]
            except Exception as e:
                raise ValueError(f"WHERE条件解析失败：{where_conditions}\n错误原因：{str(e)}") from e

    if target_col not in df_filtered.columns:
        raise KeyError(f"目标列不存在：'{target_col}'\n当前列名：{list(df_filtered.columns)}")

    unique_vals = df_filtered[target_col].unique()
    if drop_na:
        unique_vals = [val for val in unique_vals if pd.notna(val)]

    if sort_values and len(unique_vals) > 0:
        try:
            unique_vals = sorted(unique_vals)
        except TypeError:
            unique_vals = sorted([str(val) for val in unique_vals])

    result_list = list(unique_vals)
    logger.info(f"✅ 提取完成：列'{target_col}'共找到 {len(result_list)} 种不同值")
    return result_list


# ====================== 财务报表生成器 ======================
class FinancialReportGenerator:
    """护眼版财务报表生成器（适配 {月份1: {...}, 月份2: {...}} 多月份结构）"""

    def __init__(self):
        self._init_styles()
        # 核心配置参数
        self.COL_WIDTH_INDICATOR = 35  # 指标列宽度
        self.COL_WIDTH_OTHER = 20  # 总计/月份列宽度
        self.COL_WIDTH_SKU = 19  # SKU数据列宽度
        self.ROW_HEIGHT_TITLE = 40  # 标题行高度
        self.ROW_HEIGHT_HEADER = 25  # 表头行高度
        self.ROW_HEIGHT_DATA = 32  # 数据行高度
        self.ROW_HEIGHT_SPACE = 10  # 模块间空行高度

        # 结算数据业务指标显示顺序
        self.SETTLE_METRIC_ORDER = [
            "总销量", "JIT销量", "备货销量", "结算收入金额", "平台补贴金额",
            "成本金额", "退货金额", "履约售后问题金额", "仓储综合服务费", "EPR费用", "履约罚款", "其他（留空待填）",
            "结算总表利润", "每单利润", "利润率", "退货金额占比",
            "售后问题金额占比", "履约罚款率", "售后问题率总计"
        ]

        # 平台流水利指标显示顺序
        self.PLATFORM_METRIC_ORDER = [
            "卖家收入金额", "支出金额", "仓储综合服务费", "消费者及履约保障-售后问题", "EPR费用", "发货履约保障-缺货",
            "商品品质保障-质量问题", "延迟", "商品品质保障-质量问题-JIT商品", "净收入", "毛利率",
            "客单价", "转化率", "退款率", " dispute金额", "罚款金额"
        ]

    def _init_styles(self):
        """初始化Excel样式（护眼版）"""
        # 标题样式
        self.title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.title_font = Font(color="FFFFFF", bold=True, size=14)
        # 表头样式 - 结算数据
        self.settle_header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        # 表头样式 - 平台流水数据
        self.platform_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        # 表头样式 - SKC数据
        self.skc_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        # 通用样式
        self.thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.normal_font = Font(size=12)

    def _format_value(self, value: Union[int, float, str], metric: str) -> Union[str, int, float]:
        """值格式化：销量转整数、率/占比转百分比、金额保留2位小数"""
        if value is None:
            return 0

        # 处理公式标记（直接返回公式，不格式化）
        if isinstance(value, str) and value.startswith("="):
            return value

        if isinstance(value, (str, list, dict)):
            return value if "率" in metric or "占比" in metric else 0

        try:
            val = float(value)
        except (ValueError, TypeError):
            return str(value)

        if "销量" in metric:
            return int(round(val))
        elif metric == "售后履约问题率总计" or any(kw in metric for kw in ["率", "占比"]):
            return f"{val * 100:.2f}%"
        else:
            return round(val, 2)

    def _sort_metrics(self, metrics: List[str], module_type: str = "settle") -> List[str]:
        """按业务顺序排序指标"""
        if module_type == "platform":
            metric_order = self.PLATFORM_METRIC_ORDER
        else:  # 默认为结算数据
            metric_order = self.SETTLE_METRIC_ORDER

        sorted_metrics = [m for m in metric_order if m in metrics]
        extra_metrics = [m for m in metrics if m not in metric_order]
        sorted_metrics.extend(sorted(extra_metrics))
        return sorted_metrics

    def _calculate_total_value(self, metric: str, data_dict: Dict[str, Dict[str, Any]], all_months: List[str],
                               current_row: int, col_index: int = 2) -> Union[int, float, str]:
        """
        核心修改：
        1. 履约罚款/结算总表利润：仅生成Excel公式（行号固定，列号动态）
        2. 其他指标：按原有逻辑计算
        """
        # ====================== 特殊指标：固定返回Excel公式（和总计列格式完全一致） ======================
        # 列字母转换（col_index=2→B，3→C，4→D...）
        col_letter = get_column_letter(col_index)

        if metric == "履约罚款":
            # 固定公式结构：=列字母26 - 列字母28 - 列字母27 - 列字母29（如总计列=B26-B28-B27-B29，10月列=C26-C28-C27-C29）
            return f"={col_letter}29-{col_letter}30-{col_letter}31-{col_letter}32"

        elif metric == "结算总表利润":
            # 固定公式结构：=列字母7+列字母8+列字母9+列字母10+列字母11+列字母12+列字母13+列字母14+列字母15+列字母16
            return f"={col_letter}7+{col_letter}8+{col_letter}9+{col_letter}10+{col_letter}11+{col_letter}12+{col_letter}13+{col_letter}14+{col_letter}15"

        # 基础指标：总计列使用SUM公式计算所有月份之和
        base_metrics = ["总销量", "JIT销量", "备货销量", "结算收入金额", "平台补贴金额",
                        "成本金额", "退货金额", "履约售后问题金额", "仓储综合服务费", "EPR费用"]

        if metric in base_metrics:
            # 基础指标处理
            if len(all_months) == 1:
                # 只有一个月份时，直接返回该月份的原始值
                month = all_months[0]
                md = data_dict.get(month, {})
                return md.get(metric, 0)
            else:
                # 多个月份时，使用SUM公式计算总和
                start_col = get_column_letter(3)
                end_col = get_column_letter(2 + len(all_months))
                return f"=SUM({start_col}{current_row}:{end_col}{current_row})"

        # ====================== 其他指标：按原有逻辑计算 ======================
        # 先把整段时间的总基础数据算出来
        total_profit = 0
        total_income = 0
        total_sales = 0
        total_refund = 0
        total_aftersale = 0
        total_penalty = 0

        for month in all_months:
            md = data_dict.get(month, {})
            total_profit += md.get("结算总表利润", 0)
            total_income += md.get("结算收入金额", 0)
            total_sales += md.get("总销量", 0)
            total_refund += md.get("退货金额", 0)
            total_aftersale += md.get("履约售后问题金额", 0)
            # 处理履约罚款可能是字符串的情况
            penalty_val = md.get("履约罚款", 0)
            if isinstance(penalty_val, (int, float)):
                total_penalty += penalty_val

        # 按指标重新计算总计
        if metric == "利润率":
            # 返回Excel公式：利润率 = ROUND(结算总表利润 ÷ 结算收入金额 × 100, 2)%
            # 使用相对行号：结算总表利润在当前行-2，结算收入金额在当前行-9
            profit_row = current_row - 2  # 结算总表利润在当前行-2的位置
            income_row = current_row - 11  # 结算收入金额在当前行-9的位置
            return f"=ROUND(IF({col_letter}{income_row}<>0, {col_letter}{profit_row}/{col_letter}{income_row}*100, 0), 2)&\"%\""

        elif metric == "每单利润":
            # 返回Excel公式：每单利润 = ROUND(结算总表利润 ÷ 总销量, 2)
            profit_row = current_row - 1  # 结算总表利润在当前行-1的位置
            sales_row = current_row - 13  # 总销量在当前行-13的位置
            return f"=ROUND(IF({col_letter}{sales_row}<>0, {col_letter}{profit_row}/{col_letter}{sales_row}, 0), 2)"

        elif metric == "退货金额占比":
            # 总计列和月份列都使用相同的Excel公式计算
            refund_row = current_row - 9  # 退货金额在当前行-4的位置
            income_row = current_row - 12  # 结算收入金额在当前行-9的位置
            return f"=-ROUND(IF({col_letter}{income_row}<>0, {col_letter}{refund_row}/{col_letter}{income_row}*100, 0), 2)&\"%\""

        elif metric == "售后问题金额占比":
            # 总计列和月份列都使用相同的Excel公式计算
            aftersale_row = current_row - 9  # 履约售后问题金额在当前行-3的位置
            income_row = current_row - 13  # 结算收入金额在当前行-8的位置
            return f"=-ROUND(IF({col_letter}{income_row}<>0, {col_letter}{aftersale_row}/{col_letter}{income_row}*100, 0), 2)&\"%\""

        elif metric == "履约罚款率":
            # 总计列和月份列都使用相同的Excel公式计算
            penalty_row = current_row - 7  # 履约罚款在当前行-5的位置
            income_row = current_row - 14  # 结算收入金额在当前行-10的位置
            return f"=-ROUND(IF({col_letter}{income_row}<>0, {col_letter}{penalty_row}/{col_letter}{income_row}*100, 0), 2)&\"%\""

        elif metric == "售后履约问题率总计":
            # 售后履约问题率总计 = 退货金额占比 + 售后问题金额占比
            # 这两个率在当前行的上方几行，使用相对行号
            refund_rate_row = current_row - 2  # 退货金额占比在当前行-2的位置
            aftersale_rate_row = current_row - 1  # 售后问题金额占比在当前行-1的位置
            # 直接相加，因为退货金额占比和售后问题金额占比已经是百分比值
            return f"=ROUND(({col_letter}{refund_rate_row}+{col_letter}{aftersale_rate_row})*100, 2)&\"%\""

        # 其他指标：按月份求和
        total = 0
        for month in all_months:
            val = data_dict.get(month, {}).get(metric, 0)
            if isinstance(val, (int, float)):
                total += val
        return total

    def _write_data_module(
            self,
            sheet,
            current_row: int,
            module_title: str,
            data_dict: Dict[str, Dict[str, Any]],
            all_months: List[str],
            header_fill: PatternFill,
            module_type: str = "settle"
    ) -> int:
        """通用数据模块写入函数"""
        # 模块标题
        sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_TITLE
        col_num = len(all_months) + 2  # 指标 + 总计 + 月份数
        title_range = f"A{current_row}:{get_column_letter(col_num)}{current_row}"
        sheet.merge_cells(title_range)
        sheet[f"A{current_row}"] = module_title
        sheet[f"A{current_row}"].fill = self.title_fill
        sheet[f"A{current_row}"].font = self.title_font
        sheet[f"A{current_row}"].alignment = self.center_align
        current_row += 2

        # 模块表头
        sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_HEADER
        headers = ["指标", "总计"] + all_months
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
            cell.font = self.normal_font
        current_row += 1

        # 提取并排序指标
        metrics = list(set(metric for month_data in data_dict.values() for metric in month_data.keys()))
        metrics = self._sort_metrics(metrics, module_type)

        # 写入模块数据行
        for metric in metrics:
            sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_DATA
            # 写入指标
            sheet.cell(row=current_row, column=1, value=metric).alignment = self.center_align

            # 总计列（2/B列）
            total_val = self._calculate_total_value(metric, data_dict, all_months, current_row, col_index=2)
            # 公式直接写入，非公式则格式化
            sheet.cell(row=current_row, column=2).value = total_val if isinstance(total_val,
                                                                                  str) and total_val.startswith(
                "=") else self._format_value(total_val, metric)

            # 月份列（从C列/3列开始）
            for month_idx, month in enumerate(all_months, 3):
                month_val = self._calculate_total_value(metric, data_dict, [month], current_row, col_index=month_idx)
                # 公式直接写入，非公式则格式化
                sheet.cell(row=current_row, column=month_idx).value = month_val if isinstance(month_val,
                                                                                              str) and month_val.startswith(
                    "=") else self._format_value(month_val, metric)

            # 设置样式
            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.border = self.thin_border
                cell.font = self.normal_font
                cell.alignment = self.center_align
            current_row += 1

        # 模块间空行
        sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_SPACE
        current_row += 2

        return current_row

    def generate(
            self,
            raw_data: Dict[str, Dict[str, Any]],
            config: Optional[Dict[str, Any]] = None
    ) -> str:
        """生成多月份汇总财务报表"""
        config = config or {}
        output_path = config.get("output_path", "财务报表.xlsx")
        shop_name = config.get("shop_name", "未知店铺")

        # 1. 数据解析
        all_months = sorted(list(raw_data.keys()))
        if not all_months:
            raise ValueError("❌ 无有效月份数据，无法生成报表")

        # 整理结算/平台/SKC数据
        settle_data = {}
        platform_data = {}
        skc_data = {}

        for month in all_months:
            month_full_data = raw_data[month].get(month, raw_data[month])

            # 结算总表数据
            settle_month_data = month_full_data.get("结算总表数据", {})
            if settle_month_data:
                settle_data[month] = settle_month_data

            # 平台流水数据
            platform_month_data = month_full_data.get("平台流水-卖家中心", {})
            if "仓储综合服务费" in settle_month_data and "仓储综合服务费" not in platform_month_data:
                platform_month_data["仓储综合服务费"] = settle_month_data["仓储综合服务费"]
            if platform_month_data:
                platform_data[month] = platform_month_data

            # SKC数据
            skc_list = month_full_data.get("skc", [])
            for skc_item in skc_list:
                for skc_name, skc_metrics in skc_item.items():
                    if skc_name not in skc_data:
                        skc_data[skc_name] = {}
                    skc_data[skc_name][month] = skc_metrics

        # 2. 创建Excel文件
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
            sheet = writer.book.create_sheet(title="财务汇总报表")
            current_row = 1

            # 写入结算总表数据
            if settle_data:
                module_title = f"{shop_name} - 财务总表结算数据"
                current_row = self._write_data_module(
                    sheet=sheet,
                    current_row=current_row,
                    module_title=module_title,
                    data_dict=settle_data,
                    all_months=all_months,
                    header_fill=self.settle_header_fill,
                    module_type="settle"
                )

            # 写入平台流水数据
            if platform_data:
                module_title = f"平台流水-卖家中心数据"
                current_row = self._write_data_module(
                    sheet=sheet,
                    current_row=current_row,
                    module_title=module_title,
                    data_dict=platform_data,
                    all_months=all_months,
                    header_fill=self.platform_header_fill,
                    module_type="platform"
                )

            # 写入SKC数据
            # SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC SKC
            if skc_data:
                for skc_name, skc_month_data in skc_data.items():
                    # 从 SKC 数据中获取类目名
                    category_name = ""
                    for month in all_months:
                        if month in skc_month_data and "类目名" in skc_month_data[month]:
                            category_name = skc_month_data[month]["类目名"]
                            break
                    
                    # SKC模块标题
                    sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_TITLE
                    col_num = len(all_months) + 2
                    skc_title_range = f"A{current_row}:{get_column_letter(col_num)}{current_row}"
                    sheet.merge_cells(skc_title_range)
                    sheet[f"A{current_row}"] = f"SKC货号-{skc_name}-{category_name}"
                    sheet[f"A{current_row}"].fill = self.title_fill
                    sheet[f"A{current_row}"].font = self.title_font
                    sheet[f"A{current_row}"].alignment = self.center_align
                    current_row += 2

                    # SKC表头
                    sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_HEADER
                    skc_headers = ["指标", "总计"] + all_months
                    for col_idx, header in enumerate(skc_headers, 1):
                        cell = sheet.cell(row=current_row, column=col_idx, value=header)
                        cell.fill = self.skc_header_fill
                        cell.border = self.thin_border
                        cell.alignment = self.center_align
                        cell.font = self.normal_font
                    current_row += 1

                    # SKC指标排序（过滤掉"类目名"）
                    skc_metrics = list(set(
                        metric for month_data in skc_month_data.values() for metric in month_data.keys()
                        if metric != "类目名"))
                    skc_metrics = self._sort_metrics(skc_metrics, "settle")

                    # SKC数据行
                    for metric in skc_metrics:
                        sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_DATA
                        sheet.cell(row=current_row, column=1, value=metric).alignment = self.center_align

                        # 特殊处理：每单利润和利润率使用Excel公式计算
                        if metric == "每单利润":
                            # 总计列：ROUND(结算总表利润 ÷ 总销量, 2)
                            # 使用相对行号，确保每个SKC使用自己的数据
                            profit_row = current_row - 1  # 结算总表利润在当前行-1的位置
                            sales_row = current_row - 9  # 总销量在当前行-9的位置
                            sheet.cell(row=current_row, column=2,
                                       value=f"=ROUND(IF({get_column_letter(2)}{sales_row}<>0, {get_column_letter(2)}{profit_row}/{get_column_letter(2)}{sales_row}, 0), 2)")

                            # 月份列：使用ROUND公式计算
                            for month_idx, month in enumerate(all_months, 3):
                                month_col = get_column_letter(month_idx)
                                profit_row = current_row - 1  # 结算总表利润在当前行-1的位置
                                sales_row = current_row - 9  # 总销量在当前行-9的位置
                                sheet.cell(row=current_row, column=month_idx,
                                           value=f"=ROUND(IF({month_col}{sales_row}<>0, {month_col}{profit_row}/{month_col}{sales_row}, 0), 2)")
                        elif metric == "利润率":
                            # 总计列：ROUND(结算总表利润 ÷ 结算收入金额 × 100, 2)%
                            # 使用相对行号，确保每个SKC使用自己的数据
                            profit_row = current_row - 2  # 结算总表利润在当前行-2的位置
                            income_row = current_row - 7  # 结算收入金额在当前行-7的位置
                            sheet.cell(row=current_row, column=2,
                                       value=f"=ROUND(IF({get_column_letter(2)}{income_row}<>0, {get_column_letter(2)}{profit_row}/{get_column_letter(2)}{income_row}*100, 0), 2)&\"%\"")

                            # 月份列：使用ROUND公式计算
                            for month_idx, month in enumerate(all_months, 3):
                                month_col = get_column_letter(month_idx)
                                profit_row = current_row - 2  # 结算总表利润在当前行-2的位置
                                income_row = current_row - 7  # 结算收入金额在当前行-7的位置
                                sheet.cell(row=current_row, column=month_idx,
                                           value=f"=ROUND(IF({month_col}{income_row}<>0, {month_col}{profit_row}/{month_col}{income_row}*100, 0), 2)&\"%\"")
                        elif metric == "退货金额占比":
                            # 总计列：ROUND(退货金额 ÷ 结算收入金额 × 100, 2)%
                            refund_row = current_row - 5  # 退货金额在当前行-6的位置
                            income_row = current_row - 8  # 结算收入金额在当前行-7的位置
                            sheet.cell(row=current_row, column=2,
                                       value=f"=-ROUND(IF({get_column_letter(2)}{income_row}<>0, {get_column_letter(2)}{refund_row}/{get_column_letter(2)}{income_row}*100, 0), 2)&\"%\"")

                            # 月份列：使用ROUND公式计算
                            for month_idx, month in enumerate(all_months, 3):
                                month_col = get_column_letter(month_idx)
                                refund_row = current_row - 5  # 退货金额在当前行-6的位置
                                income_row = current_row - 8  # 结算收入金额在当前行-7的位置
                                sheet.cell(row=current_row, column=month_idx,
                                           value=f"=ROUND(IF({month_col}{income_row}<>0, {month_col}{refund_row}/{month_col}{income_row}*100, 0), 2)&\"%\"")

                        elif metric == "售后问题金额占比":
                            # 总计列：ROUND(履约售后问题金额 ÷ 结算收入金额 × 100, 2)%
                            aftersale_row = current_row - 5  # 履约售后问题金额在当前行-5的位置
                            income_row = current_row - 9  # 结算收入金额在当前行-7的位置
                            sheet.cell(row=current_row, column=2,
                                       value=f"=-ROUND(IF({get_column_letter(2)}{income_row}<>0, {get_column_letter(2)}{aftersale_row}/{get_column_letter(2)}{income_row}*100, 0), 2)&\"%\"")

                            # 月份列：使用ROUND公式计算
                            for month_idx, month in enumerate(all_months, 3):
                                month_col = get_column_letter(month_idx)
                                aftersale_row = current_row - 5  # 履约售后问题金额在当前行-5的位置
                                income_row = current_row - 9  # 结算收入金额在当前行-7的位置
                                sheet.cell(row=current_row, column=month_idx,
                                           value=f"=ROUND(IF({month_col}{income_row}<>0, {month_col}{aftersale_row}/{month_col}{income_row}*100, 0), 2)&\"%\"")

                        elif metric == "售后履约问题率总计":
                            # 售后履约问题率总计 = 退货金额占比 + 售后问题金额占比
                            # 这两个率在当前行的上方几行，使用相对行号
                            refund_rate_row = current_row - 2  # 退货金额占比在当前行-2的位置
                            aftersale_rate_row = current_row - 1  # 售后问题金额占比在当前行-1的位置
                            
                            # 总计列：垂直相加退货金额占比和售后问题金额占比，乘以100后取两位小数
                            sheet.cell(row=current_row, column=2,
                                       value=f"=ROUND(({get_column_letter(2)}{refund_rate_row}+{get_column_letter(2)}{aftersale_rate_row})*100, 2)&\"%\"")

                            # 月份列：垂直相加退货金额占比和售后问题金额占比，乘以100后取两位小数
                            for month_idx, month in enumerate(all_months, 3):
                                month_col = get_column_letter(month_idx)
                                sheet.cell(row=current_row, column=month_idx,
                                           value=f"=ROUND(({month_col}{refund_rate_row}+{month_col}{aftersale_rate_row})*100, 2)&\"%\"")


                        else:
                            # 其他指标：正常处理
                            # SKC总计列：使用SUM公式计算行级总和
                            if len(all_months) == 1:
                                # 只有一个月份时，直接引用月份列
                                sheet.cell(row=current_row, column=2, value=f"=C{current_row}")
                            else:
                                # 多个月份时，使用SUM范围
                                start_col = get_column_letter(3)
                                end_col = get_column_letter(2 + len(all_months))
                                sheet.cell(row=current_row, column=2,
                                           value=f"=SUM({start_col}{current_row}:{end_col}{current_row})")

                            # SKC月份列
                            for month_idx, month in enumerate(all_months, 3):
                                skc_val = skc_month_data.get(month, {}).get(metric, 0)
                                sheet.cell(row=current_row, column=month_idx, value=self._format_value(skc_val, metric))

                        # 设置样式
                        for col_idx in range(1, len(skc_headers) + 1):
                            cell = sheet.cell(row=current_row, column=col_idx)
                            cell.border = self.thin_border
                            cell.font = self.normal_font
                            cell.alignment = self.center_align
                        current_row += 1

                    # SKC模块后空行
                    sheet.row_dimensions[current_row].height = self.ROW_HEIGHT_SPACE
                    current_row += 2

            # 设置列宽
            sheet.column_dimensions[get_column_letter(1)].width = self.COL_WIDTH_INDICATOR
            col_num = len(all_months) + 2
            for col_idx in range(2, col_num + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = self.COL_WIDTH_OTHER

        # 输出结果
        abs_path = os.path.abspath(output_path)
        logger.info(f"\n✅ 多月份财务汇总报表生成成功！")
        logger.info(f"📁 文件路径：{abs_path}")
        logger.info(f"📊 报表维度：")
        logger.info(f"   - 结算总表指标数：{len(list(set(m for d in settle_data.values() for m in d.keys())))}")
        logger.info(f"   - 平台流水指标数：{len(list(set(m for d in platform_data.values() for m in d.keys())))}")
        logger.info(f"   - 月份数：{len(all_months)} | SKC组数：{len(skc_data)}")
        return abs_path

    def _generate_sku_tables(self, raw_data: Dict[str, Dict[str, Any]], config: Dict[str, Any], shop_name: str):
        """为每个月份单独生成SKU表（每个SKU一列，按SKC分组）"""
        shop_abbr = config.get("shop_abbr", "未知店铺")
        output_dir = f"配置文件_财务汇总/{shop_abbr}"
        os.makedirs(output_dir, exist_ok=True)
        
        all_months = sorted(list(raw_data.keys()))
        
        for month in all_months:
            month_full_data = raw_data[month].get(month, raw_data[month])
            sku_list = month_full_data.get("sku", [])
            
            if not sku_list:
                continue
            
            skc_sku_dict = {}
            sku_metrics_dict = {}
            skc_category_dict = {}
            for sku_item in sku_list:
                for sku_name, sku_metrics in sku_item.items():
                    sku_skc = sku_metrics.get("SKC货号", "未知SKC")
                    sku_category = sku_metrics.get("类目名", "未知类目")
                    if sku_skc not in skc_sku_dict:
                        skc_sku_dict[sku_skc] = []
                        skc_category_dict[sku_skc] = sku_category
                    skc_sku_dict[sku_skc].append(sku_name)
                    sku_metrics_dict[sku_name] = sku_metrics
            
            sku_output_path = f"{output_dir}/{month}_SKU.xlsx"
            with pd.ExcelWriter(sku_output_path, engine="openpyxl", mode="w") as writer:
                wb = writer.book
                ws = wb.create_sheet(title="SKU数据")
                
                current_row = 1
                
                ws.row_dimensions[current_row].height = self.ROW_HEIGHT_TITLE
                ws.merge_cells(f"A{current_row}:F{current_row}")
                ws[f"A{current_row}"] = f"{shop_name} - SKU数据 - {month}"
                ws[f"A{current_row}"].fill = self.title_fill
                ws[f"A{current_row}"].font = self.title_font
                ws[f"A{current_row}"].alignment = self.center_align
                current_row += 2
                
                sku_metric_order = [
                    "总销量", "JIT销量", "备货销量", "结算收入金额", "平台补贴金额",
                    "成本金额", "退货金额", "履约售后问题金额",
                    "结算总表利润", "每单利润", "利润率", "退货金额占比",
                    "售后问题金额占比", "售后履约问题率总计"
                ]
                
                for skc_name, sku_names in skc_sku_dict.items():
                    if not sku_names:
                        continue
                    
                    # 获取该SKC的原始数据（从财务汇总表的skc数据中）
                    skc_month_data = raw_data.get(month, {}).get("skc", [])
                    skc_data = {}
                    for skc_item in skc_month_data:
                        if skc_name in skc_item:
                            skc_data = skc_item[skc_name]
                            break
                    
                    ws.row_dimensions[current_row].height = self.ROW_HEIGHT_TITLE
                    col_num = len(sku_names) + 2
                    skc_title_range = f"A{current_row}:F{current_row}"
                    ws.merge_cells(skc_title_range)
                    ws[f"A{current_row}"] = f"SKC货号-{skc_name}-{skc_category_dict[skc_name]}"
                    ws[f"A{current_row}"].fill = self.title_fill
                    ws[f"A{current_row}"].font = self.title_font
                    ws[f"A{current_row}"].alignment = self.center_align
                    current_row += 2
                    
                    ws.row_dimensions[current_row].height = self.ROW_HEIGHT_HEADER
                    sku_headers = ["指标", "总计"] + sku_names
                    for col_idx, header in enumerate(sku_headers, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header)
                        cell.fill = self.skc_header_fill
                        cell.border = self.thin_border
                        cell.alignment = self.center_align
                        cell.font = self.normal_font
                    current_row += 1
                    
                    for metric in sku_metric_order:
                        if metric not in sku_metrics_dict.get(sku_names[0], {}):
                            continue
                        
                        ws.row_dimensions[current_row].height = self.ROW_HEIGHT_DATA
                        ws.cell(row=current_row, column=1, value=metric).alignment = self.center_align
                        
                        # 总计列：使用SKC原数据，不计算
                        skc_val = skc_data.get(metric, "")
                        if skc_val == "" or skc_val is None:
                            ws.cell(row=current_row, column=2, value="")
                        else:
                            ws.cell(row=current_row, column=2, value=self._format_value(skc_val, metric))
                        
                        # SKU数据列（每个SKU单独计算）
                        for sku_idx, sku_name in enumerate(sku_names, 3):
                            sku_val = sku_metrics_dict[sku_name].get(metric, 0)
                            ws.cell(row=current_row, column=sku_idx, value=self._format_value(sku_val, metric))
                        
                        for col_idx in range(1, col_num + 1):
                            cell = ws.cell(row=current_row, column=col_idx)
                            cell.border = self.thin_border
                            cell.font = self.normal_font
                            cell.alignment = self.center_align
                        current_row += 1
                    
                    ws.row_dimensions[current_row].height = self.ROW_HEIGHT_SPACE
                    current_row += 2
                    
                    ws.column_dimensions[get_column_letter(1)].width = self.COL_WIDTH_INDICATOR
                    ws.column_dimensions[get_column_letter(2)].width = self.COL_WIDTH_OTHER
                    for col_idx in range(3, col_num + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = self.COL_WIDTH_SKU
            
            logger.info(f"✅ SKU表生成成功：{sku_output_path}（{len(sku_list)}个SKU，{len(skc_sku_dict)}个SKC）")


# ====================== 数据转换工具函数 ======================
def convert_numpy_to_python(obj):
    """递归将NumPy类型转为Python原生类型"""
    if isinstance(obj, dict):
        return {k: convert_numpy_to_python(v) for k, v in obj.items()}
    elif isinstance(obj, (np.float64, np.float32)):
        return float(obj)
    elif isinstance(obj, (np.int64, np.int32, np.int_)):
        return int(obj)
    elif isinstance(obj, np.bool_):
        return bool(obj)
    elif isinstance(obj, (list, tuple)):
        return [convert_numpy_to_python(i) for i in obj]
    else:
        return obj


def make_caiwu_excel(shop_info, months_list):
    # 2. 初始化多月份数据容器
    shop_total_data = {}  # 最终整合为 {月份1: {...}, 月份2: {...}} 结构

    # 3. 逐月份提取财务数据
    for month in months_list:
        logger.info(f"\n=== 开始处理月份: {month} ===")

        # 构建文件路径（替换为你的实际文件路径）
        excel_path = f"配置文件_结算导出/{shop_info['shop_abbr']}/{month}/交易结算总表_{month}.xlsx"
        total_excel_path = f"配置文件_结算导出/{shop_info['shop_abbr']}/{month}/导出原表_卖家中心_{month}.xlsx"
        shouhou_excel_path = f"配置文件_结算导出/{shop_info['shop_abbr']}/{month}/履约保障售后问题总表_{month}.xlsx"

        try:
            # 提取当前月份数据
            month_data = start_exctract_money_excel(excel_path, total_excel_path, shouhou_excel_path, month)

            # 转换NumPy类型（避免Excel写入报错）
            clean_month_data = convert_numpy_to_python(month_data)

            # 整合到总数据容器
            shop_total_data.update(clean_month_data)

            # 打印当前月份核心数据（可选）
            # logger.info(f"\n{month} 核心数据预览：")
            # logger.info(json.dumps(clean_month_data, indent=4, ensure_ascii=False))

        except Exception as e:
            logger.error(f"❌ 处理月份 {month} 失败：{str(e)}")
            continue

    # 4. 生成多月份汇总报表
    if shop_total_data:
        # 报表配置
        report_config = {
            "output_path": f"配置文件_财务汇总/{shop_info['shop_abbr']}_财务汇总.xlsx",
            "shop_name": shop_info["shop_name"],
            "shop_abbr": shop_info["shop_abbr"]
        }

        # 初始化生成器并生成报表
        try:
            generator = FinancialReportGenerator()
            generator.generate(raw_data=shop_total_data, config=report_config)
            
            # 生成SKU表
            generator._generate_sku_tables(raw_data=shop_total_data, config=report_config, shop_name=shop_info["shop_name"])
        except PermissionError as e:
            logger.error(f"❌ 店铺{shop_info['shop_name']} 财务报表生成失败：{str(e)}")
            logger.error("❌ 财务汇总表被占用，请关闭表格后重试")

            raise Exception("❌ 财务汇总表被占用，请关闭表格后重试")

        except Exception as e:
            logger.error(f"❌ 店铺{shop_info['shop_name']} 财务报表生成失败：{str(e)}")
            raise Exception("❌ 店铺{shop_info['shop_name']} 财务报表生成失败：{str(e)}")
    else:
        logger.error("❌ 无有效财务数据，无法生成报表")
        raise Exception("❌ 无有效财务数据，无法生成报表")

# ====================== 主运行入口 ======================
if __name__ == "__main__":
    # 1. 配置参数
    uid = "111"
    months_list = ["2025.10", "2025.11", "2025.12"]  # 需要汇总的月份列表
    shop_info = get_shop_info_db(uid)

    make_caiwu_excel(shop_info, months_list)
    #
    # # 2. 初始化多月份数据容器
    # shop_total_data = {}  # 最终整合为 {月份1: {...}, 月份2: {...}} 结构
    #
    # # 3. 逐月份提取财务数据
    # for month in months_list:
    #     logger.info(f"\n=== 开始处理月份: {month} ===")
    #
    #     # 构建文件路径（替换为你的实际文件路径）
    #     excel_path = f"配置文件_结算导出/{shop_info['shop_abbr']}/{month}/交易结算总表_{month}.xlsx"
    #     total_excel_path = f"配置文件_结算导出/{shop_info['shop_abbr']}/{month}/导出原表_卖家中心_{month}.xlsx"
    #     shouhou_excel_path = f"配置文件_结算导出/{shop_info['shop_abbr']}/{month}/履约保障售后问题总表_{month}.xlsx"
    #
    #     try:
    #         # 提取当前月份数据
    #         month_data = start_exctract_money_excel(excel_path, total_excel_path, shouhou_excel_path, month)
    #
    #         # 转换NumPy类型（避免Excel写入报错）
    #         clean_month_data = convert_numpy_to_python(month_data)
    #
    #         # 整合到总数据容器
    #         shop_total_data.update(clean_month_data)
    #
    #         # 打印当前月份核心数据（可选）
    #         logger.info(f"\n{month} 核心数据预览：")
    #         logger.info(json.dumps(clean_month_data, indent=4, ensure_ascii=False))
    #
    #     except Exception as e:
    #         logger.error(f"❌ 处理月份 {month} 失败：{str(e)}")
    #         continue
    #
    # # 4. 生成多月份汇总报表
    # if shop_total_data:
    #     # 报表配置
    #     report_config = {
    #         "output_path": f"配置文件_财务汇总/{shop_info['shop_abbr']}_财务汇总.xlsx",
    #         "shop_name": shop_info["shop_name"]
    #     }
    #
    #     # 初始化生成器并生成报表
    #     generator = FinancialReportGenerator()
    #     generator.generate(raw_data=shop_total_data, config=report_config)
    # else:
    #     logger.error("❌ 无有效财务数据，无法生成报表")