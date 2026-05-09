import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from contextlib import contextmanager

from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QIcon, QColor, QBrush
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QHBoxLayout, QWidget, QPushButton,
    QHeaderView, QMessageBox, QGroupBox, QMenu, QAction, QInputDialog,
    QTabWidget, QDialog, QLabel, QComboBox, QDialogButtonBox, QLineEdit,
    QShortcut, QStatusBar, QCheckBox, QTextEdit, QSizePolicy
)
from loguru import logger

# 导入重构后的数据库操作类
from modules.classSQLite import SQLiteDB
from utils.multiThreading_log_manager import TaskStatus

# 数据库操作上下文管理器
@contextmanager
def db_connection(db_path: str):
    """数据库连接上下文管理器，自动处理连接和关闭"""
    db = None
    try:
        db = SQLiteDB(db_path)
        yield db
    except Exception as e:
        raise e
    finally:
        if db is not None:
            try:
                db.close()
            except:
                # 忽略关闭连接时的错误
                pass


class AIAnalysisWorker(QThread):
    """AI分析工作线程，异步执行AI分析任务"""
    finished = pyqtSignal(int, str)  # (record_id, result)
    error = pyqtSignal(int, str)  # (record_id, error_message)
    progress = pyqtSignal(int, str, str)  # (record_id, status_message, estimated_tokens)
    
    def __init__(self, record_id: int, prompt: str, data_content: str, ai_type: str):
        super().__init__()
        self.record_id = record_id
        self.prompt = prompt
        self.data_content = data_content
        self.ai_type = ai_type
    
    def run(self):
        """执行AI分析"""
        try:
            # 先计算预计token
            try:
                from modules.tokens_count import ai_tokens_count
                estimated_tokens = ai_tokens_count(type="text", text=self.prompt + self.data_content)[0]
            except ImportError:
                # 没有tiktoken，使用简单估算
                char_count = len(self.prompt + self.data_content)
                estimated_tokens = str(char_count // 2)
            
            # 发送进度信号，更新预计token
            self.progress.emit(self.record_id, f"预计Token: {estimated_tokens}", estimated_tokens)
            
            # 执行AI分析 - 传入原始提示词和数据内容
            from modules.AiModel import AiModel
            ai_model = AiModel("", self.data_content, self.ai_type, self.prompt)
            result = ai_model.analysis_no_stream()
            
            if result:
                self.finished.emit(self.record_id, result)
            else:
                self.error.emit(self.record_id, "AI分析返回空结果")
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.error.emit(self.record_id, f"{str(e)}\n{error_details}")


class DatabaseWorker(QThread):
    """数据库工作线程，通用化适配多表查询，支持分页和性能优化"""
    finished = pyqtSignal(list, list, int, int)  # (数据列表, 列名列表, 总记录数, 搜索结果数)
    error = pyqtSignal(str)

    def __init__(self, db_path: str, table_name: str, columns_to_display: List[str], 
                 sort_by: Tuple[str, str] = None, page: int = 1, page_size: int = 50,
                 show_subtasks: bool = False, search_conditions: List[Tuple[str, str]] = None):
        super().__init__()
        self.db_path = db_path
        self.table_name = table_name
        self.columns_to_display = columns_to_display
        self.sort_by = sort_by
        self.page = page
        self.page_size = page_size
        self.show_subtasks = show_subtasks  # 是否显示子线程任务
        self.search_conditions = search_conditions or []  # 搜索条件列表，每个元素是(字段名, 搜索值)

    def run(self):
        """线程执行入口 - 完全使用execute_sql重构，支持分页和性能优化"""
        db = None
        try:
            db = SQLiteDB(self.db_path)
            
            # 优化：设置SQLite性能参数
            db.execute_sql("PRAGMA journal_mode = WAL", fetch="none")
            db.execute_sql("PRAGMA synchronous = NORMAL", fetch="none")
            db.execute_sql("PRAGMA cache_size = 10000", fetch="none")
            db.execute_sql("PRAGMA temp_store = MEMORY", fetch="none")

            # 1. 检查表是否存在
            check_sql = """
                        SELECT name 
                        FROM sqlite_master
                        WHERE type = 'table' 
                          AND name = ? 
                        """
            table_exists = db.execute_sql(check_sql, params=(self.table_name,), fetch="fetch_one")

            if not table_exists:
                self.error.emit(f"表 '{self.table_name}' 不存在。")
                return

            # 2. 获取表结构
            info_sql = f"PRAGMA table_info({self.table_name})"
            all_table_info = db.execute_sql(info_sql, fetch="fetch")

            if not all_table_info:
                self.error.emit(f"表 '{self.table_name}' 结构为空。")
                return

            all_column_names = [col['name'] for col in all_table_info]

            # 3. 获取总记录数 - 优化：使用COUNT(id)而不是COUNT(*)提高性能
            count_sql = f"SELECT COUNT(id) as total FROM {self.table_name}"
            
            # 如果是任务表且不显示子线程任务，添加过滤条件
            if self.table_name == "task" and not self.show_subtasks:
                count_sql += " WHERE is_main_task = 1"
            
            count_result = db.execute_sql(count_sql, fetch="fetch_one")
            total_record_count = count_result.get('total', 0) if count_result else 0
            
            # 如果有搜索条件，需要获取符合条件的总记录数
            if self.search_conditions:
                # 构建WHERE子句
                where_clauses = []
                params = []
                
                # 如果是任务表且不显示子线程任务，添加过滤条件
                if self.table_name == "task" and not self.show_subtasks:
                    where_clauses.append("is_main_task = 1")
                
                # 添加搜索条件
                # 将搜索条件按搜索值分组，相同搜索值的条件用OR连接，不同搜索值用AND连接
                if self.search_conditions:
                    # 按搜索值分组
                    value_groups = {}
                    for field, value in self.search_conditions:
                        if value not in value_groups:
                            value_groups[value] = []
                        value_groups[value].append(field)
                    
                    # 为每个搜索值构建WHERE条件
                    for value, fields in value_groups.items():
                        # 支持范围搜索（如100-200）
                        if "-" in value and value.count("-") == 1:
                            try:
                                min_val, max_val = value.split("-")
                                min_val = float(min_val.strip())
                                max_val = float(max_val.strip())
                                # 范围搜索对每个字段单独处理，用OR连接
                                range_clauses = []
                                for field in fields:
                                    range_clauses.append(f"CAST({field} AS REAL) BETWEEN ? AND ?")
                                    params.extend([min_val, max_val])
                                if len(range_clauses) > 1:
                                    where_clauses.append(f"({' OR '.join(range_clauses)})")
                                else:
                                    where_clauses.append(range_clauses[0])
                            except:
                                # 如果范围解析失败，使用模糊匹配
                                like_clauses = []
                                for field in fields:
                                    like_clauses.append(f"{field} LIKE ?")
                                    params.append(f"%{value}%")
                                if len(like_clauses) > 1:
                                    where_clauses.append(f"({' OR '.join(like_clauses)})")
                                else:
                                    where_clauses.append(like_clauses[0])
                        else:
                            # 默认模糊匹配
                            like_clauses = []
                            for field in fields:
                                like_clauses.append(f"{field} LIKE ?")
                                params.append(f"%{value}%")
                            if len(like_clauses) > 1:
                                where_clauses.append(f"({' OR '.join(like_clauses)})")
                            else:
                                where_clauses.append(like_clauses[0])
                
                # 构建查询搜索结果总数的SQL
                search_count_sql = f"SELECT COUNT(id) as total FROM {self.table_name}"
                if where_clauses:
                    search_count_sql += " WHERE " + " AND ".join(where_clauses)
                
                search_count_result = db.execute_sql(search_count_sql, params=params, fetch="fetch_one")
                search_result_count = search_count_result.get('total', 0) if search_count_result else 0
            else:
                search_result_count = total_record_count

            # 4. 构建查询语句 - 使用execute_sql，支持分页
            # 如果是任务表，需要处理task_name和func_name的显示逻辑
            if self.table_name == "task":
                # 检查是否有task_name和func_name字段
                has_task_name = "task_name" in all_column_names
                has_func_name = "func_name" in all_column_names
                
                # 构建查询字段列表
                columns_to_select = []
                for col in all_column_names:
                    if col == "task_name" and has_task_name and has_func_name:
                        # 如果两个字段都存在，使用CASE WHEN来优先显示task_name
                        columns_to_select.append("CASE WHEN task_name != '' AND task_name IS NOT NULL THEN task_name ELSE func_name END as task_name")
                    elif col == "task_name" and has_func_name and not has_task_name:
                        # 如果只有func_name，将func_name重命名为task_name
                        columns_to_select.append("func_name as task_name")
                    else:
                        # 其他字段保持不变
                        columns_to_select.append(col)
                
                columns_str = ", ".join(columns_to_select)
            else:
                columns_str = ", ".join(all_column_names)
            
            sql = f"SELECT {columns_str} FROM {self.table_name}"
            
            # 构建WHERE子句
            where_clauses = []
            params = []
            
            # 如果是任务表且不显示子线程任务，添加过滤条件
            if self.table_name == "task" and not self.show_subtasks:
                where_clauses.append("is_main_task = 1")
            
            # 添加搜索条件
            # 将搜索条件按搜索值分组，相同搜索值的条件用OR连接，不同搜索值用AND连接
            if self.search_conditions:
                # 按搜索值分组
                value_groups = {}
                for field, value in self.search_conditions:
                    if value not in value_groups:
                        value_groups[value] = []
                    value_groups[value].append(field)
                
                # 为每个搜索值构建WHERE条件
                for value, fields in value_groups.items():
                    # 支持范围搜索（如100-200）
                    if "-" in value and value.count("-") == 1:
                        try:
                            min_val, max_val = value.split("-")
                            min_val = float(min_val.strip())
                            max_val = float(max_val.strip())
                            # 范围搜索对每个字段单独处理，用OR连接
                            range_clauses = []
                            for field in fields:
                                range_clauses.append(f"CAST({field} AS REAL) BETWEEN ? AND ?")
                                params.extend([min_val, max_val])
                            if len(range_clauses) > 1:
                                where_clauses.append(f"({' OR '.join(range_clauses)})")
                            else:
                                where_clauses.append(range_clauses[0])
                        except:
                            # 如果范围解析失败，使用模糊匹配
                            like_clauses = []
                            for field in fields:
                                like_clauses.append(f"{field} LIKE ?")
                                params.append(f"%{value}%")
                            if len(like_clauses) > 1:
                                where_clauses.append(f"({' OR '.join(like_clauses)})")
                            else:
                                where_clauses.append(like_clauses[0])
                    else:
                        # 默认模糊匹配
                        like_clauses = []
                        for field in fields:
                            like_clauses.append(f"{field} LIKE ?")
                            params.append(f"%{value}%")
                        if len(like_clauses) > 1:
                            where_clauses.append(f"({' OR '.join(like_clauses)})")
                        else:
                            where_clauses.append(like_clauses[0])
            
            # 添加WHERE子句到SQL
            if where_clauses:
                sql += " WHERE " + " AND ".join(where_clauses)

            if self.sort_by:
                # 安全处理排序字段，防止SQL注入
                sort_col, sort_order = self.sort_by
                if sort_col in all_column_names and sort_order.upper() in ["ASC", "DESC"]:
                    sql += f" ORDER BY {sort_col} {sort_order}"

            # 添加分页限制 - 优化：使用LIMIT和OFFSET进行分页
            offset = (self.page - 1) * self.page_size
            sql += f" LIMIT {self.page_size} OFFSET {offset}"

            # 5. 执行查询 - 使用统一的execute_sql方法
            data = db.execute_sql(sql, params=params, fetch="fetch")

            self.finished.emit(data, self.columns_to_display, total_record_count, search_result_count)

        except Exception as e:
            # 获取详细的错误信息
            import traceback
            error_details = traceback.format_exc()
            error_message = f"数据库操作失败: {str(e)}"
            
            # 记录详细错误信息
            logger.error(f"任务管理页面数据库错误: {error_message}\n{error_details}")
            
            # 发送错误信号，包含详细信息
            self.error.emit(error_message)
            
            # 如果是严重错误（如数据库连接问题），添加额外信息
            if "database" in str(e).lower() or "sqlite" in str(e).lower() or "connection" in str(e).lower():
                self.error.emit(f"数据库连接错误，请检查数据库文件是否存在且未被其他程序占用")
            
            # 如果是SQL语法错误，添加额外信息
            if "syntax" in str(e).lower() or "sql" in str(e).lower():
                self.error.emit(f"SQL查询错误，请检查表结构和查询语句")
                
        finally:
            # 确保数据库连接被正确关闭
            if db is not None:
                try:
                    db.close()
                except Exception as close_error:
                    logger.error(f"关闭数据库连接时出错: {str(close_error)}")
                    # 忽略关闭连接时的错误，不发送错误信号避免重复错误提示


class TaskStatusDialog(QDialog):
    """修改任务状态对话框"""
    def __init__(self, parent=None, task_ids: List[str] = None):
        super().__init__(parent)
        self.task_ids = task_ids or []
        self.selected_status = None
        self.init_ui()

    def init_ui(self):
        """初始化对话框UI - 使用原生样式"""
        self.setWindowTitle("修改任务状态")
        self.setWindowIcon(QIcon("gui/img/favicon.ico"))
        self.resize(380, 150)
        self.setModal(True)

        # 使用原生布局，不设置任何样式
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)


        # 提示信息（使用原生文本，不用HTML）
        info_text = f"已选择 {len(self.task_ids)} 个任务，请选择新的任务状态："
        info_label = QLabel(info_text)

        # 状态下拉框布局
        status_layout = QHBoxLayout()
        status_layout.setSpacing(10)
        status_label = QLabel("任务状态：")
        
        self.status_combo = QComboBox()
        self.status_combo.addItems([
            TaskStatus.PENDING,  # 待处理
            TaskStatus.STOPPED,  # 已退出
            TaskStatus.SUCCESS   # 已完成
        ])
        self.status_combo.setCurrentIndex(0)
        self.status_combo.setMinimumWidth(200)

        status_layout.addWidget(status_label)
        status_layout.addWidget(self.status_combo)
        status_layout.addStretch()

        # 按钮（使用原生QDialogButtonBox）
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        # 添加到布局
        layout.addWidget(info_label)
        layout.addLayout(status_layout)
        layout.addStretch()
        layout.addWidget(button_box)

    def get_selected_status(self) -> str:
        """获取选中的状态"""
        return self.status_combo.currentText()


# ===== AI分析配置 =====
# 三种表的数据列映射（不包含ID和任务ID）
AI_ANALYSIS_COLUMNS = {
    "hupu_post_list": {
        "虎扑标题": "huputitle", "虎扑分区": "hupu_zone",
        "帖子URL": "posturl", "回复数": "replies", "推荐数": "tuijian_count",
        "发帖时间": "fatietime", "添加时间": "addtime", "亮评数": "liangping_count"
    },
    "hupu_detail_list": {
        "发布内容": "fabucontent", "昵称": "nickname",
        "回复内容": "replycontent", "楼层": "floor", "IP地址": "ipaddress",
        "帖子标题": "posttitle", "点赞数": "like_count", "帖子URL": "posturl",
        "回复时间": "replytime", "添加时间": "addtime", "回复数": "reply_count"
    },
    "hupu_score_list": {
        "名称": "name", "时间": "time", "位置": "location",
        "评论": "comment", "回复评论": "reply_comment", "点赞数": "like_count",
        "评分": "score", "评分标题": "score_title", "添加时间": "addtime", "评分URL": "scoreurl"
    }
}

# 默认提示词JSON
DEFAULT_PROMPTS_JSON = """{"帖子列表": "帮我分析以下帖子列表数据，结合回复数，亮评数，推荐数综合判断热度，总结该关键词下的标题特点和情感偏向：\n数据内容:#数据内容#", "虎扑评分": "帮我分析以下虎扑评分数据，\n1.结合点赞数和用户所在地区综合判断，总结该评分对象的评分分布和用户评论的主要情\n2.如果地区发言特点明显的话总结出地区更可能出现的相关发言及情感\n数据内容:#数据内容#", "帖子详情": "帮我分析以下帖子详情回复数据：\n1.结合点赞数和用户所在IP地址综合判断，总结该帖子的用户发布内容的主要情感\n2.如果地区发言特点明显的话要特别总结出地区更可能出现的相关发言及情感，标记发言特点和地区的相关性\n3.如果用户之间有因为某些内容讨论热烈的特别罗列并分析原因\n数据内容:#数据内容#"}"""


class AIAnalysisDialog(QDialog):
    """AI分析配置对话框"""
    
    def __init__(self, parent=None, table_name: str = "", selected_rows: List[dict] = None):
        super().__init__(parent)
        self.table_name = table_name
        self.selected_rows = selected_rows or []
        self.field_checkboxes = {}
        self.init_ui()
        self.load_config()
    
    def init_ui(self):
        """初始化对话框UI"""
        # 根据表名确定任务类型名称
        type_names = {
            "hupu_post_list": "帖子列表",
            "hupu_detail_list": "帖子详情",
            "hupu_score_list": "虎扑评分"
        }
        type_name = type_names.get(self.table_name, self.table_name)
        
        self.setWindowTitle(f"AI分析配置 - {type_name}")
        self.setWindowIcon(QIcon("gui/img/favicon.ico"))
        self.resize(600, 550)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 提示信息
        info_label = QLabel(f"当前任务类型: {type_name}，已选择 {len(self.selected_rows)} 条数据")
        info_label.setStyleSheet("font-weight: bold; color: #333;")
        layout.addWidget(info_label)
        
        # 数据列选择区域
        columns_group = QGroupBox("选择要分析的数据列")
        columns_layout = QVBoxLayout(columns_group)
        
        # 全选/取消全选按钮
        select_btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("全选")
        select_all_btn.clicked.connect(self.select_all_columns)
        deselect_all_btn = QPushButton("取消全选")
        deselect_all_btn.clicked.connect(self.deselect_all_columns)
        select_btn_layout.addWidget(select_all_btn)
        select_btn_layout.addWidget(deselect_all_btn)
        select_btn_layout.addStretch()
        columns_layout.addLayout(select_btn_layout)
        
        # 复选框网格布局
        checkbox_layout = QHBoxLayout()
        checkbox_layout.setSpacing(15)
        
        columns = AI_ANALYSIS_COLUMNS.get(self.table_name, {})
        col_count = 0
        for display_name, field_name in columns.items():
            checkbox = QCheckBox(display_name)
            checkbox.setObjectName(field_name)
            self.field_checkboxes[field_name] = checkbox
            checkbox_layout.addWidget(checkbox)
            col_count += 1
            if col_count >= 4:  # 每行4个
                columns_layout.addLayout(checkbox_layout)
                checkbox_layout = QHBoxLayout()
                checkbox_layout.setSpacing(15)
                col_count = 0
        
        if col_count > 0:
            columns_layout.addLayout(checkbox_layout)
        
        layout.addWidget(columns_group)
        
        # 提示词编辑区域
        prompt_group = QGroupBox("AI提示词配置")
        prompt_layout = QVBoxLayout(prompt_group)
        
        prompt_info = QLabel("提示词中使用 #数据内容# 作为数据占位符，分析时会自动替换")
        prompt_info.setStyleSheet("color: #666; font-size: 12px;")
        prompt_layout.addWidget(prompt_info)
        
        from PyQt5.QtWidgets import QTextEdit
        self.prompt_edit = QTextEdit()
        self.prompt_edit.setPlaceholderText("请输入AI提示词...")
        self.prompt_edit.setMinimumHeight(150)
        prompt_layout.addWidget(self.prompt_edit)
        
        layout.addWidget(prompt_group)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        # 保存配置按钮
        self.save_fields_btn = QPushButton("保存列配置")
        self.save_fields_btn.clicked.connect(self.save_fields_config)
        self.save_fields_btn.setStyleSheet("background-color: #28a745; color: white;")
        
        self.save_prompt_btn = QPushButton("保存提示词")
        self.save_prompt_btn.clicked.connect(self.save_prompt_config)
        self.save_prompt_btn.setStyleSheet("background-color: #17a2b8; color: white;")
        
        # 操作按钮
        self.analyze_btn = QPushButton("开始分析")
        self.analyze_btn.clicked.connect(self.start_analysis)
        self.analyze_btn.setStyleSheet("background-color: #007BFF; color: white;")
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(self.save_fields_btn)
        btn_layout.addWidget(self.save_prompt_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.analyze_btn)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)
    
    def select_all_columns(self):
        """全选所有列"""
        for checkbox in self.field_checkboxes.values():
            checkbox.setChecked(True)
    
    def deselect_all_columns(self):
        """取消全选"""
        for checkbox in self.field_checkboxes.values():
            checkbox.setChecked(False)
    
    def load_config(self):
        """加载配置"""
        from config.common_config import config_manager
        
        # 默认字段配置
        default_fields = {
            "hupu_detail_list": ["fabucontent", "replycontent", "ipaddress", "posttitle", "like_count", "reply_count"],
            "hupu_post_list": ["huputitle", "hupu_zone", "replies", "tuijian_count", "liangping_count"],
            "hupu_score_list": ["name", "location", "comment", "reply_comment", "like_count", "score", "score_title", "scoreurl"]
        }
        
        # 加载数据列配置
        fields_key = f"AIAnalysis_{self.table_name}_fields"
        saved_fields = config_manager.get_or_set_config(fields_key, "")
        if saved_fields:
            try:
                selected_fields = json.loads(saved_fields) if isinstance(saved_fields, str) else saved_fields
                for field_name in selected_fields:
                    if field_name in self.field_checkboxes:
                        self.field_checkboxes[field_name].setChecked(True)
            except:
                # 解析失败，使用默认字段配置
                default_fields_list = default_fields.get(self.table_name, [])
                for field_name in default_fields_list:
                    if field_name in self.field_checkboxes:
                        self.field_checkboxes[field_name].setChecked(True)
        else:
            # 使用默认字段配置，而不是全选
            default_fields_list = default_fields.get(self.table_name, [])
            for field_name in default_fields_list:
                if field_name in self.field_checkboxes:
                    self.field_checkboxes[field_name].setChecked(True)
        
        # 加载提示词配置 - 从SettingPage_ai_speech_content读取JSON
        type_names = {
            "hupu_post_list": "帖子列表",
            "hupu_detail_list": "帖子详情",
            "hupu_score_list": "虎扑评分"
        }
        type_name = type_names.get(self.table_name)
        
        if type_name:
            # 从SettingPage_ai_speech_content读取提示词JSON
            speech_config = config_manager.get_or_set_config("SettingPage_ai_speech_content", DEFAULT_PROMPTS_JSON)
            try:
                speech_dict = json.loads(speech_config) if speech_config else {}
                saved_prompt = speech_dict.get(type_name, "")
            except:
                saved_prompt = ""
            
            # 如果没有配置，从默认JSON中读取
            if not saved_prompt:
                try:
                    default_dict = json.loads(DEFAULT_PROMPTS_JSON)
                    saved_prompt = default_dict.get(type_name, "")
                except:
                    saved_prompt = ""
            
            self.prompt_edit.setPlainText(saved_prompt)
    
    def save_fields_config(self):
        """保存数据列配置"""
        from config.common_config import config_manager
        
        selected_fields = []
        for field_name, checkbox in self.field_checkboxes.items():
            if checkbox.isChecked():
                selected_fields.append(field_name)
        
        fields_key = f"AIAnalysis_{self.table_name}_fields"
        config_manager.upsert_config(fields_key, json.dumps(selected_fields, ensure_ascii=False))
        
        QMessageBox.information(self, "成功", "数据列配置已保存")
    
    def save_prompt_config(self):
        """保存提示词配置到JSON"""
        from config.common_config import config_manager
        
        prompt = self.prompt_edit.toPlainText().strip()
        if not prompt:
            QMessageBox.warning(self, "警告", "提示词不能为空")
            return
        
        # 获取类型名称
        type_names = {
            "hupu_post_list": "帖子列表",
            "hupu_detail_list": "帖子详情",
            "hupu_score_list": "虎扑评分"
        }
        type_name = type_names.get(self.table_name)
        
        if type_name:
            # 获取现有提示词JSON配置
            speech_config = config_manager.get_or_set_config("SettingPage_ai_speech_content", DEFAULT_PROMPTS_JSON)
            try:
                speech_dict = json.loads(speech_config) if speech_config else {}
            except:
                speech_dict = {}
            
            # 更新对应类型的提示词
            speech_dict[type_name] = prompt
            config_manager.upsert_config("SettingPage_ai_speech_content", 
                                         json.dumps(speech_dict, ensure_ascii=False))
        
        QMessageBox.information(self, "成功", "提示词配置已保存到JSON配置")
    
    def get_selected_fields(self) -> List[str]:
        """获取选中的字段列表"""
        selected_fields = []
        for field_name, checkbox in self.field_checkboxes.items():
            if checkbox.isChecked():
                selected_fields.append(field_name)
        return selected_fields
    
    def get_prompt(self) -> str:
        """获取提示词"""
        return self.prompt_edit.toPlainText().strip()
    
    def start_analysis(self):
        """开始分析"""
        selected_fields = self.get_selected_fields()
        if not selected_fields:
            QMessageBox.warning(self, "警告", "请至少选择一个数据列")
            return
        
        prompt = self.get_prompt()
        if not prompt:
            QMessageBox.warning(self, "警告", "请输入AI提示词")
            return
        
        if not self.selected_rows:
            QMessageBox.warning(self, "警告", "没有选中任何数据")
            return
        
        # 接受对话框，由父组件执行分析
        self.accept()


class ExportOrderDialog(QDialog):
    """导出订单数据对话框"""
    
    def __init__(self, parent=None, task_id=None, table_name=None, data=None):
        super().__init__(parent)
        self.task_id = task_id
        self.table_name = table_name
        self.data = data  # 用于导出选中行的数据
        self.init_ui()
    
    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("导出数据")
        self.setWindowIcon(QIcon("gui/img/favicon.ico"))
        self.resize(450, 350)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)
        
        # 标题
        title_label = QLabel("导出数据")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #333;")
        title_layout = QHBoxLayout()
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        layout.addLayout(title_layout)
        
        # 提示信息
        if self.data:
            # 导出选中行
            info_text = f"当前选中 {len(self.data)} 条数据，表名: {self.table_name}"
            detail_label = QLabel("注意：将导出当前选中的数据")
        else:
            # 导出任务数据
            info_text = f"当前任务: {self.table_name}，任务ID: {self.task_id}"
            detail_label = QLabel("注意：将导出该任务ID对应的所有数据，不仅仅是当前页面显示的数据")
        
        info_label = QLabel(info_text)
        info_label.setStyleSheet("font-weight: bold; color: #555; margin: 10px 0;")
        layout.addWidget(info_label)
        
        # 添加详细说明
        detail_label.setStyleSheet("color: #888; font-size: 12px; margin: 5px 0;")
        layout.addWidget(detail_label)
        
        # 导出格式选择
        format_group = QGroupBox("选择导出格式")
        format_layout = QVBoxLayout(format_group)
        
        self.format_combo = QComboBox()
        self.format_combo.addItems(["Excel (.xlsx)", "CSV (.csv)", "文本文件 (.txt)"])
        self.format_combo.setCurrentIndex(0)  # 默认选择Excel
        format_layout.addWidget(self.format_combo)
        
        layout.addWidget(format_group)
        
        # 文件名设置
        file_group = QGroupBox("文件名设置")
        file_layout = QVBoxLayout(file_group)
        
        # 自动生成文件名
        auto_name_layout = QHBoxLayout()
        auto_name_label = QLabel("自动生成文件名:")
        self.auto_name_checkbox = QCheckBox()
        self.auto_name_checkbox.setChecked(True)
        self.auto_name_checkbox.stateChanged.connect(self.on_auto_name_changed)
        auto_name_layout.addWidget(auto_name_label)
        auto_name_layout.addWidget(self.auto_name_checkbox)
        auto_name_layout.addStretch()
        file_layout.addLayout(auto_name_layout)
        
        # 文件名输入
        filename_layout = QHBoxLayout()
        filename_label = QLabel("文件名:")
        self.filename_edit = QLineEdit()
        self.filename_edit.setEnabled(False)  # 默认禁用，使用自动生成
        self.filename_edit.setPlaceholderText("例如: 订单数据_20231225")
        filename_layout.addWidget(filename_label)
        filename_layout.addWidget(self.filename_edit)
        file_layout.addLayout(filename_layout)
        
        layout.addWidget(file_group)
        
        # 导出选项
        options_group = QGroupBox("导出选项")
        options_layout = QVBoxLayout(options_group)
        
        self.include_headers_checkbox = QCheckBox("包含表头")
        self.include_headers_checkbox.setChecked(True)
        options_layout.addWidget(self.include_headers_checkbox)
        
        layout.addWidget(options_group)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        self.export_btn = QPushButton("导出")
        self.export_btn.clicked.connect(self.export_data)
        self.export_btn.setStyleSheet("background-color: #007BFF; color: white; font-weight: bold;")
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)
    
    def on_auto_name_changed(self, state):
        """自动生成文件名复选框状态改变"""
        if state == Qt.Checked:
            self.filename_edit.setEnabled(False)
            self.filename_edit.clear()
        else:
            self.filename_edit.setEnabled(True)
    
    def get_export_format(self):
        """获取导出格式"""
        format_text = self.format_combo.currentText()
        if "Excel" in format_text:
            return "excel"
        elif "CSV" in format_text:
            return "csv"
        else:
            return "txt"
    
    def get_filename(self):
        """获取文件名"""
        if self.auto_name_checkbox.isChecked():
            # 自动生成文件名
            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            if self.data:
                # 导出选中行
                return f"{self.table_name}_选中数据_{now}"
            else:
                # 导出任务数据
                return f"{self.table_name}_任务{self.task_id}_{now}"
        else:
            # 使用用户输入的文件名
            return self.filename_edit.text().strip()
    
    def export_data(self):
        """导出数据"""
        # 不检查数据，因为数据将在用户确认后获取
        filename = self.get_filename()
        if not filename:
            QMessageBox.warning(self, "警告", "请输入文件名")
            return
        
        # 接受对话框，由父组件执行导出
        self.accept()


class CustomTableWidgetItem(QTableWidgetItem):
    """自定义表格项，支持正确的整数排序和状态颜色"""

    def __init__(self, text: str, sort_key=None, status_color=None):
        super().__init__(text)
        self.sort_key = sort_key if sort_key is not None else text
        self.status_color = status_color

    def __lt__(self, other):
        """重载排序比较逻辑"""
        try:
            return int(self.sort_key) < int(other.sort_key)
        except (ValueError, TypeError):
            return super().__lt__(other)


class TableTabWidget(QWidget):
    """单个分页标签页的内容组件，包含表格和操作按钮"""

    def __init__(self, parent, db_path: str, table_name: str,
                 columns_to_display: List[str], column_aliases: Dict[str, str],
                 column_width_config: Dict[str, Tuple[str, int]],
                 context_menu_actions: List[Dict[str, callable]],
                 tab_name: str = ""):
        super().__init__(parent)
        # 基础配置
        self.db_path = db_path
        self.table_name = table_name
        self.columns_to_display = columns_to_display
        self.column_aliases = column_aliases
        self.column_width_config = column_width_config
        self.context_menu_actions = context_menu_actions
        self.tab_name = tab_name

        # 运行时状态
        self.current_sort = ("id", "ASC")
        self.all_data = []
        self.original_column_names = []
        
        # AI分析工作线程列表（支持多个并行任务）
        self.ai_workers = []
        
        # 交互防抖/重入控制
        self.worker = None
        self._is_loading_data = False
        self._pending_load_page = None
        self._context_menu_open = False
        self._action_locks = {}
        self._action_last_trigger_ms = {}
        
        # 分页相关状态
        self.page_size = 50  # 每页显示50条数据
        self.current_page = 1  # 当前页码
        self.total_pages = 1  # 总页数
        self.total_count = 0  # 总记录数

        # 初始化UI
        self.init_ui()
    
    def __del__(self):
        """析构函数，确保清理工作线程"""
        try:
            if hasattr(self, 'worker') and self.worker and self.worker.isRunning():
                self.worker.terminate()
                self.worker.wait()
        except:
            pass

    def init_ui(self):
        """初始化标签页UI - 将按钮和表格放在GroupBox中"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        # 创建GroupBox，标题为分页名字
        group_box = QGroupBox(self.tab_name if self.tab_name else "数据管理")
        group_layout = QVBoxLayout(group_box)
        group_layout.setContentsMargins(10, 15, 10, 30)

        # 顶部按钮布局
        top_layout = QHBoxLayout()

        self.refresh_btn = QPushButton("刷新")
        self.refresh_btn.setFixedWidth(100)
        self.refresh_btn.setIcon(QIcon("./gui/img/shuaxin.png"))
        self.refresh_btn.clicked.connect(self.load_data)

        self.sort_btn = QPushButton("排序")
        self.sort_btn.setFixedWidth(100)
        self.sort_btn.setIcon(QIcon("./gui/img/paixu.png"))
        self.sort_btn.clicked.connect(self.toggle_sort)

        top_layout.addWidget(self.refresh_btn)
        top_layout.addWidget(self.sort_btn)
        
        # 添加筛选功能
        self.add_search_controls(top_layout)
        
        # 如果是虎扑评分表，添加打开虎扑评分首页按钮
        if 'score' in self.table_name.lower():
            self.open_hupu_score_btn = QPushButton("打开虎扑评分首页")
            self.open_hupu_score_btn.setIcon(QIcon("./gui/img/tijiao.png"))
            self.open_hupu_score_btn.clicked.connect(self.open_hupu_score_homepage)
            top_layout.addWidget(self.open_hupu_score_btn)
        
        # 如果是帖子列表或帖子详情表，添加打开帖子搜索首页按钮
        if 'post' in self.table_name.lower():
            self.open_hupu_search_btn = QPushButton("打开帖子搜索首页")
            self.open_hupu_search_btn.setIcon(QIcon("./gui/img/tijiao.png"))
            self.open_hupu_search_btn.clicked.connect(self.open_hupu_post_search_homepage)
            top_layout.addWidget(self.open_hupu_search_btn)
        
        top_layout.addStretch()

        # 表格组件 - 保持原有样式
        self.table_widget = QTableWidget()
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setSortingEnabled(False)
        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_widget.customContextMenuRequested.connect(self.show_context_menu)
        self.table_widget.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        
        # 设置表头可交互，支持拉伸
        self.table_widget.horizontalHeader().setHighlightSections(False)  # 点击表头不高亮
        self.table_widget.horizontalHeader().setSectionsMovable(True)  # 允许拖动列顺序
        self.table_widget.horizontalHeader().setStretchLastSection(True)  # 最后一列拉伸填充剩余空间
        
        # 设置表头右键菜单
        self.table_widget.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_widget.horizontalHeader().customContextMenuRequested.connect(self.show_header_context_menu)
        
        # 使用QSS样式文件
        from .static.qss import load_table_style
        self.table_widget.setStyleSheet(load_table_style())
        
        # 设置选择模式为整行选择，支持多选
        self.table_widget.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        
        # 设置行高控制，防止行被拉长
        self.table_widget.verticalHeader().setDefaultSectionSize(30)  # 默认行高30像素
        self.table_widget.verticalHeader().setMinimumSectionSize(25)  # 最小行高25像素
        self.table_widget.verticalHeader().setMaximumSectionSize(100)  # 最大行高100像素
        self.table_widget.setWordWrap(False)  # 禁用自动换行
        
        # 使用QSS样式文件
        from .static.qss import load_table_style
        self.table_widget.setStyleSheet(load_table_style())

        # 分页控件布局
        pagination_layout = QHBoxLayout()
        
        # 页码输入框 - 改为普通输入框
        self.page_input = QLineEdit()
        self.page_input.setFixedWidth(80)
        self.page_input.returnPressed.connect(self.go_to_page)  # 按回车键也可以跳转
        
        # 前往按钮
        self.go_to_page_btn = QPushButton("前往")
        self.go_to_page_btn.setFixedWidth(60)
        self.go_to_page_btn.clicked.connect(self.go_to_page)
        
        # 上一页按钮
        self.prev_page_btn = QPushButton("上一页")
        self.prev_page_btn.setFixedWidth(80)
        self.prev_page_btn.clicked.connect(self.go_to_prev_page)
        
        # 下一页按钮
        self.next_page_btn = QPushButton("下一页")
        self.next_page_btn.setFixedWidth(80)
        self.next_page_btn.clicked.connect(self.go_to_next_page)
        
        # 分页大小选择
        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["20", "50", "100", "200"])
        self.page_size_combo.setCurrentText(str(self.page_size))
        self.page_size_combo.setFixedWidth(80)
        self.page_size_combo.currentTextChanged.connect(self.on_page_size_changed)
        
        # 左侧区域
        left_layout = QHBoxLayout()
        left_layout.addWidget(QLabel("每页显示"))
        left_layout.addWidget(self.page_size_combo)
        left_layout.addWidget(QLabel("条"))
        left_layout.addWidget(QLabel(" | "))
        self.current_data_info_label = QLabel("共0条数据，当前页0条")
        left_layout.addWidget(self.current_data_info_label)
        
        # 右侧区域
        right_layout = QHBoxLayout()
        self.page_info_label = QLabel("共0页，当前第0页")
        right_layout.addWidget(self.page_info_label)
        right_layout.addWidget(QLabel(" | "))
        right_layout.addWidget(QLabel("跳转到"))
        right_layout.addWidget(self.page_input)
        right_layout.addWidget(self.go_to_page_btn)
        right_layout.addWidget(self.prev_page_btn)
        right_layout.addWidget(self.next_page_btn)
        
        # 添加到主布局
        pagination_layout.addLayout(left_layout)
        pagination_layout.addStretch()
        pagination_layout.addLayout(right_layout)

        # 将按钮和表格添加到GroupBox布局中
        group_layout.addLayout(top_layout)
        group_layout.addWidget(self.table_widget)
        group_layout.addLayout(pagination_layout)

        # 将GroupBox添加到主布局
        layout.addWidget(group_box)
        
        # 添加快捷键支持
        self.copy_shortcut = QShortcut(Qt.CTRL + Qt.Key_C, self.table_widget)
        self.copy_shortcut.activated.connect(self.copy_selected_rows)
        
        # 初始加载数据
        QTimer.singleShot(100, self.load_data)

    def add_search_controls(self, layout):
        """添加搜索控件"""
        # 第一个搜索条件
        self.search_combo1 = QComboBox()
        self.search_combo1.addItems(["全部"])
        self.search_combo1.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        
        self.search_input1 = QLineEdit()
        self.search_input1.setPlaceholderText("搜索条件1")
        self.search_input1.setFixedWidth(150)
        self.search_input1.setToolTip("支持模糊搜索，数字支持范围输入，如100-200")
        
        # 第二个搜索条件
        self.search_combo2 = QComboBox()
        self.search_combo2.addItems(["全部"])
        self.search_combo2.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        
        self.search_input2 = QLineEdit()
        self.search_input2.setPlaceholderText("搜索条件2")
        self.search_input2.setFixedWidth(150)
        self.search_input2.setToolTip("支持模糊搜索，数字支持范围输入，如100-200")
        
        # 搜索按钮
        self.search_btn = QPushButton("搜索")
        self.search_btn.setIcon(QIcon("./gui/img/sousuo.png"))
        self.search_btn.clicked.connect(self.apply_search)
        
        # 清除搜索按钮
        self.clear_search_btn = QPushButton("清除")
        self.clear_search_btn.setIcon(QIcon("./gui/img/qingli.png"))
        self.clear_search_btn.clicked.connect(self.clear_search)
        
        # 如果是任务管理页面，添加显示子线程任务复选框
        self.show_subtasks_checkbox = None
        if self.table_name == "task":
            from PyQt5.QtWidgets import QCheckBox
            self.show_subtasks_checkbox = QCheckBox("显示子线程任务")
            self.show_subtasks_checkbox.setToolTip("勾选后显示子线程任务，不勾选只显示主线程任务")
            self.show_subtasks_checkbox.setChecked(False)  # 默认不勾选，只显示主线程任务
            self.show_subtasks_checkbox.stateChanged.connect(self.on_show_subtasks_changed)
        
        # 添加到布局
        layout.addWidget(QLabel("搜索:"))
        layout.addWidget(self.search_combo1)
        layout.addWidget(self.search_input1)
        layout.addWidget(self.search_combo2)
        layout.addWidget(self.search_input2)
        layout.addWidget(self.search_btn)
        layout.addWidget(self.clear_search_btn)
        
        # 如果是任务管理页面，添加显示子线程任务复选框
        if self.show_subtasks_checkbox:
            layout.addWidget(self.show_subtasks_checkbox)
        
        # 初始化搜索选项
        self.init_search_options()
        
        # 保存原始数据引用
        self.original_data = []
        self.filtered_data = []

    def on_show_subtasks_changed(self, state):
        """显示子线程任务复选框状态改变时的处理"""
        # 刷新数据
        self.refresh_data()
    
    def init_search_options(self):
        """初始化搜索选项"""
        # 根据不同的表设置不同的搜索选项
        if self.table_name == "ai_analysis":
            options = [
                "全部", "ID", "任务名称", "状态", "信息", "备注", "任务ID", "类型", "AI总结"
            ]
            self.field_mapping = {
                "全部": self.columns_to_display,
                "ID": "id", "任务名称": "task_name", "状态": "status", "信息": "msg", "备注": "remarks", 
                "任务ID": "task_id", "类型": "type", "AI总结": "ai_sumup"
            }
        elif self.table_name == "hupu_post_list":
            options = [
                "全部", "ID", "虎扑标题", "虎扑分区", "帖子URL", "回复数",
                "推荐数", "发帖时间", "添加时间", "亮评数", "任务ID"
            ]
            self.field_mapping = {
                "全部": self.columns_to_display,
                "ID": "id", "虎扑标题": "huputitle", "虎扑分区": "hupu_zone", 
                "帖子URL": "posturl", "回复数": "replies", "推荐数": "tuijian_count",
                "发帖时间": "fatietime", "添加时间": "addtime", "亮评数": "liangping_count",
                "任务ID": "task_id"  # 将订单ID改为任务ID
            }
        elif self.table_name == "hupu_detail_list":
            options = [
                "全部", "ID", "发布内容", "昵称", "回复内容", "楼层",
                "IP地址", "帖子标题", "点赞数", "帖子URL", "回复时间",
                "添加时间", "任务ID", "回复数"
            ]
            self.field_mapping = {
                "全部": self.columns_to_display,
                "ID": "id", "发布内容": "fabucontent", "昵称": "nickname", 
                "回复内容": "replycontent", "楼层": "floor", "IP地址": "ipaddress",
                "帖子标题": "posttitle", "点赞数": "like_count", "帖子URL": "posturl",
                "回复时间": "replytime", "添加时间": "addtime", "任务ID": "task_id",  # 将订单ID改为任务ID
                "回复数": "reply_count"
            }
        elif self.table_name == "hupu_score_list":
            options = [
                "全部", "ID", "名称", "时间", "位置", "评论", "回复评论",
                "点赞数", "评分", "评分标题", "添加时间", "任务ID", "评分URL"
            ]
            self.field_mapping = {
                "全部": self.columns_to_display,
                "ID": "id", "名称": "name", "时间": "time", "位置": "location",
                "评论": "comment", "回复评论": "reply_comment", "点赞数": "like_count",
                "评分": "score", "评分标题": "score_title", "添加时间": "addtime",
                "任务ID": "task_id",  # 将订单ID改为任务ID
                "评分URL": "scoreurl"
            }
        else:
            # 默认选项（任务管理、店铺管理等）
            if self.table_name == "task":
                # 任务管理的中文选项
                options = [
                    "全部", "ID", "任务名称", "状态", "函数名称", "任务组", "信息", "备注", "任务ID", 
                    "代理IP", "创建时间", "更新时间"
                ]
                self.field_mapping = {
                    "全部": self.columns_to_display,
                    "ID": "id", "任务名称": "task_name", "状态": "status", "函数名称": "func_name",
                    "任务组": "task_group", "信息": "msg", "备注": "remarks", "任务ID": "task_id", "代理IP": "ip",
                    "创建时间": "create_time", "更新时间": "update_time"
                }
            elif self.table_name == "shops":
                # 店铺管理的中文选项
                options = [
                    "全部", "ID", "店铺名称", "缩写", "手机号", "密码",
                    "连接状态", "创建时间", "更新时间", "请求头", "Cookies"
                ]
                self.field_mapping = {
                    "全部": self.columns_to_display,
                    "ID": "id", "店铺名称": "shop_name", "缩写": "shop_abbr", "手机号": "phone",
                    "密码": "password", "连接状态": "connect_status", "创建时间": "create_time",
                    "更新时间": "update_time", "请求头": "headers", "Cookies": "cookies"
                }
            else:
                # 其他表的默认选项
                options = ["全部"] + self.columns_to_display
                self.field_mapping = {"全部": self.columns_to_display}
                for col in self.columns_to_display:
                    self.field_mapping[col] = col
        
        # 更新下拉框选项
        self.search_combo1.clear()
        self.search_combo1.addItems(options)
        self.search_combo2.clear()
        self.search_combo2.addItems(options)

    def apply_search(self):
        """应用搜索条件 - 直接从数据库查询符合条件的数据"""
        # 重置到第一页
        self.current_page = 1
        # 调用load_data，它会传递搜索条件并从数据库查询
        self.load_data(page=1)

    def match_condition(self, cell_value, search_value):
        """检查单元格值是否匹配搜索条件"""
        # 支持范围搜索（如100-200）
        if "-" in search_value and search_value.count("-") == 1:
            try:
                min_val, max_val = search_value.split("-")
                min_val = float(min_val.strip())
                max_val = float(max_val.strip())
                cell_num = float(cell_value)
                return min_val <= cell_num <= max_val
            except:
                pass
        
        # 默认模糊匹配
        return search_value.lower() in cell_value.lower()

    def clear_search(self):
        """清除搜索条件"""
        self.search_input1.clear()
        self.search_input2.clear()
        # 重置搜索状态
        self.is_searching = False
        # 重置到第一页并重新加载数据
        self.current_page = 1
        self.load_data(page=1)

    def update_table_with_searched_data(self):
        """使用搜索后的数据更新表格"""
        # 这个方法现在主要用于刷新显示，实际数据已经在populate_table中设置
        # 由于搜索现在在数据库层面进行，所以不需要额外的过滤
        pass

    def load_data(self, page: int = None):
        """通用数据加载方法 - 支持分页和加载提示"""
        if page is None:
            page = self.current_page

        # 如果正在加载，则只记录最后一次请求，避免频繁点击导致重入
        if self._is_loading_data:
            self._pending_load_page = page
            return

        self._is_loading_data = True
        self._pending_load_page = None

        self.refresh_btn.setEnabled(False)
        self.table_widget.clearContents()

        # 显示加载提示
        self.show_loading_indicator()

        # 获取是否显示子线程任务的状态
        show_subtasks = False
        if hasattr(self, 'show_subtasks_checkbox') and self.show_subtasks_checkbox:
            show_subtasks = self.show_subtasks_checkbox.isChecked()

        # 获取搜索条件
        search_conditions = []
        field1 = self.search_combo1.currentText()
        value1 = self.search_input1.text().strip()
        field2 = self.search_combo2.currentText()
        value2 = self.search_input2.text().strip()

        # 构建搜索条件列表
        if value1:
            if field1 == "全部":
                # 选择"全部"时，对所有列进行搜索
                columns = self.field_mapping.get("全部", self.columns_to_display)
                for col in columns:
                    search_conditions.append((col, value1))
            else:
                # 选择特定列时，只搜索该列
                db_field = self.field_mapping.get(field1, field1)
                search_conditions.append((db_field, value1))

        if value2:
            if field2 == "全部":
                # 选择"全部"时，对所有列进行搜索
                columns = self.field_mapping.get("全部", self.columns_to_display)
                for col in columns:
                    search_conditions.append((col, value2))
            else:
                # 选择特定列时，只搜索该列
                db_field = self.field_mapping.get(field2, field2)
                search_conditions.append((db_field, value2))

        # 设置搜索状态标志
        self.is_searching = len(search_conditions) > 0

        # 启动数据库线程
        worker = DatabaseWorker(
            self.db_path, self.table_name,
            self.columns_to_display, self.current_sort, page, self.page_size,
            show_subtasks=show_subtasks, search_conditions=search_conditions
        )
        self.worker = worker
        worker.finished.connect(self.populate_table)
        worker.error.connect(self.show_error_with_details)

        # 添加错误处理，防止线程启动失败
        try:
            worker.start()
        except Exception as e:
            self.hide_loading_indicator()
            self.refresh_btn.setEnabled(True)
            self._is_loading_data = False

            import traceback
            error_details = traceback.format_exc()
            error_message = f"启动数据加载线程失败: {str(e)}"
            logger.error(f"任务管理页面启动线程错误: {error_message}\n{error_details}")
            QMessageBox.critical(self, "错误", f"{error_message}\n\n详细信息已记录到日志文件")

    
    def get_ai_analysis_results(self):
        """获取AI分析结果 - 简化版本，避免卡死"""
        try:
            # 获取选中行的任务ID
            selected_ids = self.get_selected_task_ids()
            if not selected_ids:
                QMessageBox.information(self, "提示", "请选择要查看AI分析数据的任务")
                return
            
            task_id = selected_ids[0]  # 使用第一个选中的任务ID
            
            # 获取任务类型
            task_type = self.get_task_type_by_id(task_id)
            if not task_type:
                QMessageBox.warning(self, "警告", "无法确定任务类型")
                return
            
            # 获取父级DbTableViewer实例
            parent_viewer = self.get_parent_viewer()
            if not parent_viewer:
                QMessageBox.warning(self, "警告", "无法访问标签页控件")
                return
            
            # 查找ai_analysis标签页
            ai_analysis_tab = None
            for i in range(parent_viewer.tab_widget.count()):
                tab_widget = parent_viewer.tab_widget.widget(i)
                if hasattr(tab_widget, 'table_name') and tab_widget.table_name == "ai_analysis":
                    ai_analysis_tab = tab_widget
                    break
            
            if not ai_analysis_tab:
                QMessageBox.warning(self, "警告", "找不到AI分析结果标签页")
                return
            
            # 切换到AI分析标签页
            parent_viewer.tab_widget.setCurrentWidget(ai_analysis_tab)
            
            # 直接设置搜索条件并刷新数据，不使用QTimer延迟
            if hasattr(ai_analysis_tab, 'search_combo1') and hasattr(ai_analysis_tab, 'search_input1'):
                # 设置筛选条件为任务ID
                index = ai_analysis_tab.search_combo1.findText("任务ID")
                if index >= 0:
                    ai_analysis_tab.search_combo1.setCurrentIndex(index)
                    ai_analysis_tab.search_input1.setText(task_id)
                    ai_analysis_tab.search_input2.clear()
                    
                    # 直接刷新数据，不使用apply_search，避免分页问题
                    if hasattr(ai_analysis_tab, 'load_data'):
                        ai_analysis_tab.load_data()
                    elif hasattr(ai_analysis_tab, 'refresh_data'):
                        ai_analysis_tab.refresh_data()
            else:
                # 如果没有搜索功能，尝试直接刷新
                if hasattr(ai_analysis_tab, 'load_data'):
                    ai_analysis_tab.load_data()
                elif hasattr(ai_analysis_tab, 'refresh_data'):
                    ai_analysis_tab.refresh_data()
            
            QMessageBox.information(self, "成功", f"已切换到AI分析结果页面，任务ID: {task_id}")
                
        except Exception as e:
            # 获取详细的错误信息
            import traceback
            error_details = traceback.format_exc()
            error_message = f"获取AI分析数据失败: {str(e)}"
            
            # 记录详细错误信息
            logger.error(f"获取AI分析数据错误: {error_message}\n{error_details}")
            
            # 显示错误对话框
            QMessageBox.critical(self, "错误", f"{error_message}\n\n详细信息已记录到日志文件")

    def refresh_ai_analysis_tab(self, ai_analysis_tab, task_id):
        """刷新AI分析标签页数据"""
        try:
            # 设置搜索条件并刷新数据 - 参考open_related_data_by_task_type的实现
            if hasattr(ai_analysis_tab, 'search_combo1') and hasattr(ai_analysis_tab, 'search_input1'):
                # 设置筛选条件为任务ID
                index = ai_analysis_tab.search_combo1.findText("任务ID")
                if index >= 0:
                    ai_analysis_tab.search_combo1.setCurrentIndex(index)
                    ai_analysis_tab.search_input1.setText(task_id)
                    ai_analysis_tab.search_input2.clear()
                    
                    # 直接刷新数据，不使用apply_search，避免分页问题
                    if hasattr(ai_analysis_tab, 'load_data'):
                        ai_analysis_tab.load_data()
                    elif hasattr(ai_analysis_tab, 'refresh_data'):
                        ai_analysis_tab.refresh_data()
            else:
                # 如果没有搜索功能，尝试直接刷新
                if hasattr(ai_analysis_tab, 'load_data'):
                    ai_analysis_tab.load_data()
                elif hasattr(ai_analysis_tab, 'refresh_data'):
                    ai_analysis_tab.refresh_data()
        except Exception as e:
            print(f"刷新AI分析标签页失败: {str(e)}")

    def populate_table(self, data: list, column_names: list, total_record_count: int, search_result_count: int):
        """填充表格数据 - 支持分页和大数据量优化"""
        # 隐藏加载指示器
        self.hide_loading_indicator()
        
        # 保存数据
        self.original_data = data.copy()
        self.filtered_data = data.copy()  # 保持兼容性，但现在与original_data相同
        
        self.all_data = data
        self.original_column_names = column_names.copy()
        
        # 保存搜索结果数和总记录数
        self.total_record_count = total_record_count
        self.search_result_count = search_result_count
        self.total_count = search_result_count  # 兼容性，用于分页计算
        
        # 初始化搜索状态（如果没有设置，默认为False）
        if not hasattr(self, 'is_searching'):
            self.is_searching = search_result_count < total_record_count
        
        # 计算总页数（基于搜索结果数）
        self.total_pages = (self.total_count + self.page_size - 1) // self.page_size
        
        # 设置表格行列数
        self.table_widget.setRowCount(len(data))
        self.table_widget.setColumnCount(len(column_names))

        # 设置表头（使用别名）
        display_headers = [self.column_aliases.get(col, col) for col in column_names]
        self.table_widget.setHorizontalHeaderLabels(display_headers)

        # 优化大数据量渲染：暂停UI更新，批量设置数据
        self.table_widget.setUpdatesEnabled(False)
        
        try:
            # 填充单元格数据
            for row_idx, row_data in enumerate(data):
                for col_idx, col_name in enumerate(column_names):
                    cell_value = row_data.get(col_name, "")
                    display_text = self.format_cell_value(cell_value)

                    # 判断是否是状态列，设置颜色
                    status_colors = None
                    if col_name == "status":
                        status_colors = self.get_status_color(cell_value)

                    # 创建自定义表格项
                    item = CustomTableWidgetItem(display_text, cell_value, status_colors)
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    
                    # 如果有颜色，应用前景色和背景色
                    if status_colors:
                        foreground_color, background_color = status_colors
                        item.setForeground(QBrush(QColor(foreground_color)))
                        item.setBackground(QBrush(QColor(background_color)))
                    
                    self.table_widget.setItem(row_idx, col_idx, item)
        finally:
            # 恢复UI更新
            self.table_widget.setUpdatesEnabled(True)

        # 设置列宽
        self.setup_column_widths()
        
        # 如果设置了保持列宽标志，应用自定义列宽
        if hasattr(self, '_keep_column_widths') and self._keep_column_widths and hasattr(self, '_column_widths'):
            # 使用定时器延迟应用，确保表格完全渲染后再设置列宽
            QTimer.singleShot(50, self._apply_custom_column_widths)
        
        # 更新分页信息
        self.update_pagination_info()

        # 更新UI状态
        self.refresh_btn.setEnabled(True)
        self.update_sort_button_text()
        
        # 加载完成收尾（解锁并处理排队请求）
        self._finish_data_load()


    def format_cell_value(self, value) -> str:
        """格式化单元格值的显示 - 限制显示长度，防止行高增加"""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, (dict, list)):
            # JSON数据限制显示长度
            json_str = json.dumps(value, indent=0, ensure_ascii=False)
            if len(json_str) > 100:
                return json_str[:100] + "..."
            return json_str
        elif isinstance(value, str):
            # 字符串限制显示长度
            if len(value) > 200:
                return value[:200] + "..."
            return value
        elif value is None:
            return ""
        else:
            # 其他类型转为字符串后限制长度
            str_value = str(value)
            if len(str_value) > 200:
                return str_value[:200] + "..."
            return str_value

    def get_status_color(self, status: str) -> Optional[Tuple[str, str]]:
        """根据任务状态返回对应的颜色（与前端保持一致）
        返回: (前景色, 背景色) 或 None
        """
        if not status:
            return None
        
        status = str(status).lower()
        
        # 状态颜色映射（与前端CSS保持一致）
        status_color_map = {
            # 成功状态 - 绿色
            '已完成': ('#0f7a3b', '#e8fff2'),
            'success': ('#0f7a3b', '#e8fff2'),
            '执行成功': ('#0f7a3b', '#e8fff2'),
            
            # 失败状态 - 红色
            '异常': ('#b91c1c', '#fff0f0'),
            'failed': ('#b91c1c', '#fff0f0'),
            '执行失败': ('#b91c1c', '#fff0f0'),
            '已超时': ('#b91c1c', '#fff0f0'),
            'timeout': ('#b91c1c', '#fff0f0'),
            '已退出': ('#b91c1c', '#fff0f0'),
            'stopped': ('#b91c1c', '#fff0f0'),
            
            # 进行中状态 - 橙色/黄色
            '进行中': ('#a15c00', '#fff8e1'),
            'running': ('#a15c00', '#fff8e1'),
            '执行中': ('#a15c00', '#fff8e1'),
            
            # 待处理状态 - 蓝色
            '待处理': ('#125f9c', '#e7f3ff'),
            'pending': ('#125f9c', '#e7f3ff'),
            '待执行': ('#125f9c', '#e7f3ff'),
            
            # 验证码状态 - 紫色
            '验证码': ('#5e07a8', '#faf5ff'),
            'captcha': ('#5e07a8', '#faf5ff'),
        }
        
        return status_color_map.get(status)

    def setup_column_widths(self):
        """设置列宽 - 支持左右拉伸"""
        header = self.table_widget.horizontalHeader()
        
        # 设置所有列为可拉伸模式，但不是固定宽度
        for col_idx, col_name in enumerate(self.original_column_names):
            # 获取列宽配置，用于设置最小宽度
            config = self.column_width_config.get(col_name, ("Stretch", 100))
            resize_mode, width = config
            
            # 设置为Interactive模式，允许用户通过拖动调整宽度
            header.setSectionResizeMode(col_idx, QHeaderView.Interactive)
            
            # 设置初始宽度和最小宽度
            if width > 0:
                self.table_widget.setColumnWidth(col_idx, width)
                header.setMinimumSectionSize(width // 2)  # 最小宽度为初始宽度的一半
            else:
                self.table_widget.setColumnWidth(col_idx, 100)  # 默认宽度
                header.setMinimumSectionSize(50)  # 最小宽度为50像素
            
            # 设置最大宽度，防止某一列过宽
            header.setMaximumSectionSize(500)
        
        # 设置最后一列自动拉伸填充剩余空间
        if self.original_column_names:
            last_col_idx = len(self.original_column_names) - 1
            header.setSectionResizeMode(last_col_idx, QHeaderView.Stretch)

    def show_error(self, message: str):
        """显示错误提示 - 保持原有逻辑"""
        self.hide_loading_indicator()
        QMessageBox.critical(self, "错误", message)
        self.refresh_btn.setEnabled(True)
        self._finish_data_load()
    
    def show_error_with_details(self, message: str, details: str = None):
        """显示详细错误信息，便于调试"""
        self.hide_loading_indicator()
        self.refresh_btn.setEnabled(True)
        self._finish_data_load()
            
        # 如果有详细信息，记录到日志
        if details:
            import traceback
            logger.error(f"任务管理页面错误详情: {message}\n{details}")
            print(f"任务管理页面错误详情:\n{message}\n{details}")
            traceback.print_exc()
        
        # 显示错误对话框
        QMessageBox.critical(self, "错误", f"{message}\n\n详细信息已记录到日志文件")

    def _finish_data_load(self):
        """结束一次数据加载，并在有排队请求时自动触发下一次加载"""
        self._is_loading_data = False
        self.worker = None

        if self._pending_load_page is not None:
            next_page = self._pending_load_page
            self._pending_load_page = None
            QTimer.singleShot(0, lambda p=next_page: self.load_data(p))

    
    def show_loading_indicator(self):
        """显示加载指示器"""
        # 如果已经有加载指示器，先隐藏
        self.hide_loading_indicator()
        
        # 直接在表格上显示加载信息
        self.table_widget.setRowCount(1)
        self.table_widget.setColumnCount(1)
        
        # 创建加载标签项
        loading_item = QTableWidgetItem("正在加载数据，请稍候...")
        loading_item.setTextAlignment(Qt.AlignCenter)
        loading_item.setFlags(Qt.ItemIsEnabled)  # 禁止选择
        
        # 设置到表格中
        self.table_widget.setItem(0, 0, loading_item)
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        # 移除verticalHeader().setStretchLastSection(True)，避免行高被拉伸
    
    def hide_loading_indicator(self):
        """隐藏加载指示器"""
        # 不需要特殊处理，因为populate_table会重新设置表格内容
        pass

    def toggle_sort(self):
        """切换排序方式 - 支持分页"""
        col, order = self.current_sort
        new_order = "DESC" if order == "ASC" else "ASC"
        self.current_sort = (col, new_order)
        # 排序时重置到第一页
        self.current_page = 1
        self.load_data()

    def update_sort_button_text(self):
        """更新排序按钮文本 - 保持原有逻辑"""
        self.sort_btn.setText(f"排序")

    def on_header_clicked(self, logical_index: int):
        """表头点击排序 - 支持分页"""
        if logical_index >= len(self.original_column_names):
            return

        clicked_col = self.original_column_names[logical_index]
        if clicked_col == self.current_sort[0]:
            self.toggle_sort()
        else:
            self.current_sort = (clicked_col, "ASC")
            # 排序时重置到第一页
            self.current_page = 1
            self.load_data()
        
        # 如果设置了保持列宽标志，重新应用自定义列宽
        if hasattr(self, '_keep_column_widths') and self._keep_column_widths and hasattr(self, '_column_widths'):
            if self._column_widths:
                # 使用定时器延迟应用，确保数据加载完成后再设置列宽
                QTimer.singleShot(100, self._apply_custom_column_widths)
    
    def _apply_custom_column_widths(self):
        """应用自定义列宽设置"""
        try:
            # 检查必要的属性是否存在
            if not hasattr(self, '_column_widths') or not self._column_widths:
                return
                
            if not hasattr(self, 'table_widget') or self.table_widget is None:
                return
                
            # 禁用自动拉伸
            self.table_widget.horizontalHeader().setStretchLastSection(False)
            
            # 恢复列宽
            for column, width in self._column_widths.items():
                if column < self.table_widget.columnCount():  # 确保列索引有效
                    self.table_widget.setColumnWidth(column, width)
                    
            # 重新启用自动拉伸，但只对最后一列生效
            self.table_widget.horizontalHeader().setStretchLastSection(True)
        except Exception as e:
            print(f"应用自定义列宽失败: {str(e)}")
            import traceback
            traceback.print_exc()

    def _release_action_lock(self, action_key: str):
        self._action_locks[action_key] = False

    def _invoke_with_guard(self, action_key: str, callback, *args, cooldown_ms: int = 500):
        """统一的动作防抖+重入保护，避免频繁点击导致卡死/崩溃"""
        now_ms = int(datetime.now().timestamp() * 1000)
        last_ms = self._action_last_trigger_ms.get(action_key, 0)

        if self._action_locks.get(action_key, False):
            return
        if now_ms - last_ms < cooldown_ms:
            return

        self._action_locks[action_key] = True
        self._action_last_trigger_ms[action_key] = now_ms

        def _run_action():
            try:
                callback(*args)
            except Exception as e:
                logger.error(f"执行动作失败[{action_key}]: {str(e)}")
            finally:
                QTimer.singleShot(cooldown_ms, lambda k=action_key: self._release_action_lock(k))

        QTimer.singleShot(0, _run_action)

    def show_context_menu(self, position):
        """显示右键菜单 - 支持task表和其他表的不同ID获取方式"""
        if self._context_menu_open:
            return

        self._context_menu_open = True
        try:
            menu = QMenu(self)
            
            # 获取点击位置的单元格
            item = self.table_widget.itemAt(position)
            
            # 添加URL相关的菜单项
            self.add_url_menu_items(menu, position)
            
            # 添加复制单元格选项（无论是否有选中行）
            if item:
                copy_cell_action = QAction("复制单元格", self)
                copy_cell_action.triggered.connect(lambda checked=False, c=item: self._invoke_with_guard("copy_cell", self.copy_cell_content, c, cooldown_ms=200))
                menu.addAction(copy_cell_action)
                menu.addSeparator()

            # 统一使用数据库索引id（所有表都使用get_selected_row_ids）
            selected_ids = self.get_selected_row_ids()
            if not selected_ids:
                menu.addAction("无选中行")
                menu.exec_(self.table_widget.mapToGlobal(position))
                return

            # 添加复制选中行选项
            copy_rows_action = QAction("复制选中行 (Ctrl+C)", self)
            copy_rows_action.triggered.connect(lambda checked=False: self._invoke_with_guard("copy_rows", self.copy_selected_rows, cooldown_ms=200))
            menu.addAction(copy_rows_action)
            
            # 如果是任务表，添加联合搜索选项
            if self.table_name == "task":
                # 检查是否有spider权限且是否是虎扑相关任务
                task_ids = self.get_selected_task_ids()
                is_hupu_task = False
                task_type = None
                task_id = ""
                
                if task_ids:
                    task_id = task_ids[0]
                    task_type = self.get_task_type_by_id(task_id)
                    # 检查返回的任务类型是否是虎扑相关
                    if task_type in ["帖子列表", "帖子详情", "虎扑评分"]:
                        is_hupu_task = True
                
                # 只有spider权限且是虎扑任务时才显示"打开相关数据"
                if self.check_spider_permission() and is_hupu_task:
                    menu.addSeparator()
                    open_related_action = QAction("打开相关数据", self)
                    open_related_action.triggered.connect(lambda checked=False: self._invoke_with_guard("open_related_data", self.open_related_data_by_task_type))
                    menu.addAction(open_related_action)
                
                # 如果是虎扑相关任务，添加AI分析和导出选项
                if is_hupu_task:
                        # AI分析菜单项（放在AI分析结果上面）
                        ai_action = QAction("AI分析", self)
                        ai_action.triggered.connect(lambda checked=False, tid=task_id, ttype=task_type: self._invoke_with_guard(f"ai_analysis_{tid}", self.show_ai_analysis_from_task, tid, ttype, cooldown_ms=800))
                        menu.addAction(ai_action)
                        
                        # AI分析结果菜单项
                        ai_analysis_action = QAction("AI分析结果", self)
                        ai_analysis_action.triggered.connect(lambda checked=False: self._invoke_with_guard("ai_analysis_results", self.get_ai_analysis_results, cooldown_ms=800))
                        menu.addAction(ai_analysis_action)
                        
                        # 导出订单数据菜单项（只有爬虫权限才可使用）
                        export_order_action = QAction("导出订单数据", self)
                        export_order_action.triggered.connect(lambda checked=False, tid=task_id, ttype=task_type: self._invoke_with_guard(f"export_order_{tid}", self.export_task_order_data, tid, ttype, cooldown_ms=800))
                        menu.addAction(export_order_action)
            
            # 如果是AI分析结果表，添加查看分析结果选项
            if self.table_name == "ai_analysis":
                menu.addSeparator()
                view_result_action = QAction("查看分析结果", self)
                view_result_action.triggered.connect(lambda checked=False: self._invoke_with_guard("view_ai_analysis_result", self.view_ai_analysis_result, cooldown_ms=600))
                menu.addAction(view_result_action)
            
            menu.addSeparator()

            # 添加配置的菜单项
            for action_config in self.context_menu_actions:
                action = QAction(list(action_config.keys())[0], self)
                action_name = list(action_config.keys())[0]
                action.triggered.connect(lambda checked=False, f=list(action_config.values())[0], ids=selected_ids, n=action_name: self._invoke_with_guard(f"menu_{n}", f, ids))
                menu.addAction(action)

            menu.exec_(self.table_widget.mapToGlobal(position))
        except Exception as e:
            # 获取详细的错误信息
            import traceback
            error_details = traceback.format_exc()
            error_message = f"显示右键菜单时出错: {str(e)}"
            
            # 记录详细错误信息
            logger.error(f"任务管理页面右键菜单错误: {error_message}\n{error_details}")
            
            # 显示错误对话框
            QMessageBox.critical(self, "错误", f"{error_message}\n\n详细信息已记录到日志文件")
            
            # 仍然显示基本菜单，避免完全无法使用
            try:
                menu = QMenu(self)
                basic_action = QAction("刷新页面", self)
                basic_action.triggered.connect(lambda: self.load_data())
                menu.addAction(basic_action)
                menu.exec_(self.table_widget.mapToGlobal(position))
            except:
                pass
        finally:
            self._context_menu_open = False

    def get_selected_row_ids(self) -> List[int]:
        """获取选中行的ID列表 - 保持原有逻辑"""
        selected_rows = set(item.row() for item in self.table_widget.selectedItems())
        ids = []

        for row in selected_rows:
            id_item = self.table_widget.item(row, 0)
            if id_item and isinstance(id_item, CustomTableWidgetItem):
                try:
                    ids.append(int(id_item.sort_key))
                except (ValueError, TypeError):
                    continue

        return ids

    def show_ai_analysis_from_task(self, task_id: str, task_type: str):
        """从任务管理表发起AI分析"""
        try:
            # 根据任务类型确定表名
            table_mapping = {
                "帖子列表": "hupu_post_list",
                "帖子详情": "hupu_detail_list",
                "虎扑评分": "hupu_score_list"
            }
            table_name = table_mapping.get(task_type)
            if not table_name:
                QMessageBox.warning(self, "警告", f"不支持的任务类型: {task_type}")
                return
            
            # 从虎扑数据库获取该任务ID的数据
            from config.common_config import hupu_db
            if hupu_db is None:
                QMessageBox.warning(self, "警告", "虎扑数据库未初始化")
                return
            
            query = f"SELECT * FROM {table_name} WHERE task_id = ?"
            rows = hupu_db.execute_sql(query, params=(task_id,), fetch="fetch")
            
            if not rows:
                QMessageBox.information(self, "提示", f"任务ID {task_id} 没有找到相关数据")
                return
            
            # 显示对话框
            dialog = AIAnalysisDialog(self, table_name, rows)
            if dialog.exec_() == QDialog.Accepted:
                self.execute_ai_analysis_for_task(dialog, rows, table_name)
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"AI分析对话框错误: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"AI分析失败: {str(e)}")
    
    def execute_ai_analysis_for_task(self, dialog: 'AIAnalysisDialog', rows_data: List[dict], table_name: str):
        """从任务执行AI分析 - 异步执行"""
        try:
            selected_fields = dialog.get_selected_fields()
            prompt = dialog.get_prompt()
            
            # 获取字段映射
            columns = AI_ANALYSIS_COLUMNS.get(table_name, {})
            reverse_mapping = {v: k for k, v in columns.items()}
            
            # 获取任务ID（从第一条数据）
            task_id = rows_data[0].get("task_id", "") if rows_data else ""
            
            # 构建数据内容
            data_content_parts = []
            for i, row in enumerate(rows_data, 1):
                row_parts = []
                for field_name in selected_fields:
                    display_name = reverse_mapping.get(field_name, field_name)
                    value = row.get(field_name, "")
                    if value:
                        row_parts.append(f"{display_name}: {value}")
                if row_parts:
                    data_content_parts.append(f"【第{i}条】\n" + "\n".join(row_parts))
            
            data_content = "\n\n".join(data_content_parts)
            
            # 替换提示词中的占位符
            final_prompt = prompt.replace("#数据内容#", data_content)
            
            # 确定AI类型
            type_mapping = {
                "hupu_post_list": "post_list",
                "hupu_detail_list": "detail_list",
                "hupu_score_list": "score_list"
            }
            ai_type = type_mapping.get(table_name, "")
            
            # 确定任务类型名称
            type_name_mapping = {
                "hupu_post_list": "帖子列表",
                "hupu_detail_list": "帖子详情",
                "hupu_score_list": "虎扑评分"
            }
            task_type = type_name_mapping.get(table_name, "")
            
            # 先插入"进行中"记录（token在线程中计算）
            record_id = self.insert_ai_analysis_pending(task_id, task_type, len(rows_data))
            if record_id is None:
                QMessageBox.warning(self, "警告", "无法创建AI分析记录")
                return
            
            # 刷新AI分析结果标签页
            self.refresh_ai_analysis_tab()
            
            # 创建并启动异步工作线程 - 传入原始提示词和数据内容
            worker = AIAnalysisWorker(record_id, prompt, data_content, ai_type)
            worker.finished.connect(self.on_ai_analysis_finished)
            worker.error.connect(self.on_ai_analysis_error)
            worker.finished.connect(lambda _rid, _res, w=worker: self.cleanup_ai_worker(w))
            worker.error.connect(lambda _rid, _err, w=worker: self.cleanup_ai_worker(w))
            worker.progress.connect(self.on_ai_analysis_progress)
            
            # 添加到列表并启动
            self.ai_workers.append(worker)
            worker.start()
            
            # 提示用户分析已开始
            QMessageBox.information(self, "提示", "AI分析任务已开始执行\n请到AI分析结果表查看进度")
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"AI分析执行错误: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"AI分析执行失败: {str(e)}")
    
    def insert_ai_analysis_pending(self, task_id: str, task_type: str, data_count: int) -> Optional[int]:
        """插入一条"进行中"状态的AI分析记录，返回记录ID"""
        try:
            from config.common_config import hupu_db, get_current_time
            
            if hupu_db is None:
                logger.warning("虎扑数据库未初始化")
                return None
            
            current_time = get_current_time()
            
            insert_sql = """
                INSERT INTO ai_analysis (task_name, status, msg, remarks, task_id, type, ai_sumup)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """
            params = (
                f"{task_type}AI分析",  # task_name
                "进行中",  # status
                f"分析中...预计消耗tokens 计算中（共 {data_count} 条数据）",  # msg
                f"开始时间: {current_time}\n预计Token: 计算中...",  # remarks
                task_id,  # task_id
                task_type,  # type
                ""  # ai_sumup 暂时为空
            )
            
            hupu_db.execute_sql(insert_sql, params=params, fetch="none")
            
            # 获取最后插入的ID
            result = hupu_db.execute_sql("SELECT last_insert_rowid() as id", fetch="fetch_one")
            record_id = result.get("id") if result else None
            
            logger.info(f"已创建AI分析记录，ID: {record_id}，状态: 进行中")
            return record_id
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"插入AI分析记录失败: {str(e)}\n{error_details}")
            return None
    
    def cleanup_ai_worker(self, worker):
        """清理已完成的工作线程"""
        try:
            if worker in self.ai_workers:
                self.ai_workers.remove(worker)
            worker.deleteLater()
        except Exception as e:
            logger.error(f"清理AI工作线程失败: {str(e)}")
    
    def on_ai_analysis_progress(self, record_id: int, status_message: str, estimated_tokens: str):
        """AI分析进度回调 - token计算完成后更新备注和msg"""
        try:
            from config.common_config import hupu_db
            if hupu_db is None:
                return
            
            # 更新备注和msg字段
            update_sql = """
                UPDATE ai_analysis
                SET remarks = REPLACE(remarks, '预计Token: 计算中...', ?),
                    msg = ?
                WHERE id = ?
            """
            hupu_db.execute_sql(update_sql, params=(
                f"预计Token: {estimated_tokens}",
                f"分析中...预计消耗tokens {estimated_tokens}",
                record_id
            ), fetch="none")
            
            # 刷新标签页
            self.refresh_ai_analysis_tab()
            
        except Exception as e:
            logger.error(f"更新Token进度失败: {str(e)}")
    
    def on_ai_analysis_finished(self, record_id: int, result: str):
        """AI分析完成回调"""
        try:
            from config.common_config import hupu_db, get_current_time
            
            if hupu_db is None:
                logger.warning("虎扑数据库未初始化")
                return
            
            current_time = get_current_time()
            
            # 更新记录状态和结果
            update_sql = """
                UPDATE ai_analysis 
                SET status = ?, msg = ?, remarks = remarks || ?, ai_sumup = ?
                WHERE id = ?
            """
            params = (
                "已完成",
                "分析完成",
                f"\n完成时间: {current_time}",
                result,
                record_id
            )
            
            hupu_db.execute_sql(update_sql, params=params, fetch="none")
            logger.info(f"AI分析记录已更新，ID: {record_id}，状态: 已完成")
            
            # 刷新AI分析结果标签页
            self.refresh_ai_analysis_tab()
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"更新AI分析记录失败: {str(e)}\n{error_details}")
    
    def on_ai_analysis_error(self, record_id: int, error_message: str):
        """AI分析错误回调"""
        try:
            from config.common_config import hupu_db, get_current_time
            
            if hupu_db is None:
                logger.warning("虎扑数据库未初始化")
                return
            
            current_time = get_current_time()
            
            # 更新记录状态为失败
            update_sql = """
                UPDATE ai_analysis 
                SET status = ?, msg = ?, remarks = remarks || ?
                WHERE id = ?
            """
            params = (
                "失败",
                f"分析失败: {error_message[:100]}",
                f"\n失败时间: {current_time}",
                record_id
            )
            
            hupu_db.execute_sql(update_sql, params=params, fetch="none")
            logger.error(f"AI分析记录失败，ID: {record_id}，错误: {error_message}")
            
            # 刷新AI分析结果标签页
            self.refresh_ai_analysis_tab()
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"更新AI分析错误记录失败: {str(e)}\n{error_details}")
    
    def refresh_ai_analysis_tab(self):
        """刷新AI分析结果标签页"""
        try:
            # 查找AI分析结果标签页并刷新
            parent_viewer = self.parent()
            if parent_viewer and hasattr(parent_viewer, 'tab_widget'):
                for i in range(parent_viewer.tab_widget.count()):
                    tab = parent_viewer.tab_widget.widget(i)
                    if hasattr(tab, 'table_name') and tab.table_name == "ai_analysis":
                        if hasattr(tab, 'load_data'):
                            tab.load_data()
                        break
        except Exception as e:
            logger.error(f"刷新AI分析结果标签页失败: {str(e)}")

    def view_ai_analysis_result(self):
        """查看AI分析结果 - 从ai_analysis表右键菜单调用"""
        try:
            # 获取选中行的ID
            selected_ids = self.get_selected_row_ids()
            if not selected_ids:
                QMessageBox.information(self, "提示", "请选择要查看的记录")
                return
            
            record_id = selected_ids[0]
            
            # 从数据库获取该记录的分析结果
            from config.common_config import hupu_db
            if hupu_db is None:
                QMessageBox.warning(self, "警告", "虎扑数据库未初始化")
                return
            
            query = "SELECT ai_sumup, status, task_name FROM ai_analysis WHERE id = ?"
            result = hupu_db.execute_sql(query, params=(record_id,), fetch="fetch_one")
            
            if not result:
                QMessageBox.warning(self, "警告", "未找到该记录")
                return
            
            ai_sumup = result.get("ai_sumup", "")
            status = result.get("status", "")
            task_name = result.get("task_name", "AI分析结果")
            
            # 显示结果
            if ai_sumup and ai_sumup.strip():
                self.show_ai_result_dialog(ai_sumup)
            else:
                status_text = status if status else "未开始"
                QMessageBox.information(self, "提示", f"该记录暂无分析结果\n当前状态: {status_text}")
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"查看AI分析结果失败: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"查看分析结果失败: {str(e)}")

    def show_ai_analysis_dialog(self):
        """显示AI分析配置对话框并执行分析"""
        try:
            # 获取选中行的数据
            selected_rows = set(item.row() for item in self.table_widget.selectedItems())
            if not selected_rows:
                QMessageBox.information(self, "提示", "请选择要分析的数据行")
                return
            
            # 收集选中行的数据
            rows_data = []
            for row in selected_rows:
                row_dict = {}
                for col in range(self.table_widget.columnCount()):
                    header = self.table_widget.horizontalHeaderItem(col)
                    item = self.table_widget.item(row, col)
                    if header and item:
                        # 使用列别名作为键
                        col_name = self.column_aliases.get(header.text(), header.text())
                        row_dict[col_name] = item.text()
                rows_data.append(row_dict)
            
            # 显示对话框
            dialog = AIAnalysisDialog(self, self.table_name, rows_data)
            if dialog.exec_() == QDialog.Accepted:
                # 用户确认分析
                self.execute_ai_analysis(dialog, rows_data)
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"AI分析对话框错误: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"AI分析失败: {str(e)}")
    
    def execute_ai_analysis(self, dialog: AIAnalysisDialog, rows_data: List[dict]):
        """执行AI分析"""
        try:
            selected_fields = dialog.get_selected_fields()
            prompt = dialog.get_prompt()
            
            # 获取字段映射（显示名 -> 字段名）
            columns = AI_ANALYSIS_COLUMNS.get(self.table_name, {})
            reverse_mapping = {v: k for k, v in columns.items()}  # 字段名 -> 显示名
            
            # 构建数据内容
            data_content_parts = []
            for i, row in enumerate(rows_data, 1):
                row_parts = []
                for field_name in selected_fields:
                    display_name = reverse_mapping.get(field_name, field_name)
                    # 尝试从行数据中获取值（可能用的是显示名或字段名）
                    value = row.get(display_name) or row.get(field_name, "")
                    if value:
                        row_parts.append(f"{display_name}: {value}")
                if row_parts:
                    data_content_parts.append(f"【第{i}条】\n" + "\n".join(row_parts))
            
            data_content = "\n\n".join(data_content_parts)
            
            # 替换提示词中的占位符
            final_prompt = prompt.replace("#数据内容#", data_content)
            
            # 确定AI类型
            type_mapping = {
                "hupu_post_list": "post_list",
                "hupu_detail_list": "detail_list",
                "hupu_score_list": "score_list"
            }
            ai_type = type_mapping.get(self.table_name, "")
            
            # 调用AI模型 - 传入用户临时修改的提示词
            from modules.AiModel import AiModel
            ai_model = AiModel("", data_content, ai_type, prompt)
            result = ai_model.analysis_no_stream()
            
            # 显示结果
            if result:
                self.show_ai_result_dialog(result)
            else:
                QMessageBox.warning(self, "警告", "AI分析返回空结果")
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"AI分析执行错误: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"AI分析执行失败: {str(e)}")
    
    _ai_result_dialog = None  # 类变量，防止重复打开对话框
    
    def show_ai_result_dialog(self, result: str):
        """显示AI分析结果对话框"""
        # 防止重复打开
        if self._ai_result_dialog is not None and self._ai_result_dialog.isVisible():
            self._ai_result_dialog.raise_()
            self._ai_result_dialog.activateWindow()
            return
        
        from PyQt5.QtWidgets import QTextEdit, QSizePolicy
        
        dialog = QDialog(self)
        dialog.setWindowTitle("AI分析结果")
        dialog.setWindowIcon(QIcon("gui/img/favicon.ico"))
        dialog.resize(600, 500)
        dialog.setAttribute(Qt.WA_DeleteOnClose)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # 结果文本框
        result_edit = QTextEdit()
        result_edit.setReadOnly(True)
        result_edit.setPlainText(result)
        result_edit.setMinimumHeight(400)
        layout.addWidget(result_edit)
        
        # 按钮
        btn_layout = QHBoxLayout()
        
        copy_btn = QPushButton("复制结果")
        copy_btn.setIcon(QIcon("gui/img/fuzhi.png"))
        copy_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        copy_btn.clicked.connect(lambda: QApplication.clipboard().setText(result))
        
        close_btn = QPushButton("关闭")
        close_btn.setIcon(QIcon("gui/img/close.png"))
        close_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        close_btn.clicked.connect(dialog.accept)
        
        btn_layout.addWidget(copy_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
        
        # 清理引用
        dialog.destroyed.connect(self._clear_ai_result_dialog)
        self._ai_result_dialog = dialog
        dialog.show()
    
    def _clear_ai_result_dialog(self):
        """清理对话框引用"""
        self._ai_result_dialog = None

    def open_related_order(self):
        """打开相关订单"""
        try:
            # 获取选中行的数据
            selected_rows = set(item.row() for item in self.table_widget.selectedItems())
            if not selected_rows:
                QMessageBox.information(self, "提示", "请选择要查看订单的行")
                return
            
            # 获取第一行的任务ID
            first_row = min(selected_rows)
            taskid_item = self.table_widget.item(first_row, self.get_column_index_by_name("task_id"))
            if not taskid_item:
                QMessageBox.warning(self, "警告", "未找到任务ID")
                return
            
            taskid = taskid_item.text().strip()
            if not taskid:
                QMessageBox.warning(self, "警告", "任务ID为空")
                return
            
            # 切换到订单列表标签页
            parent_viewer = self.get_parent_viewer()  # 使用更可靠的方式获取DbTableViewer实例
            if parent_viewer:
                # 查找订单列表标签页
                for i in range(parent_viewer.tab_widget.count()):
                    tab_text = parent_viewer.tab_widget.tabText(i)
                    if "订单列表" in tab_text:
                        parent_viewer.tab_widget.setCurrentIndex(i)
                        order_tab = parent_viewer.tab_widget.currentWidget()
                        
                        # 等待标签页加载完成
                        QTimer.singleShot(500, lambda: self.filter_order_by_taskid(order_tab, taskid))
                        break
                else:
                    QMessageBox.warning(self, "警告", "未找到订单列表标签页")
            else:
                QMessageBox.warning(self, "警告", "无法访问标签页控件")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"打开订单失败: {str(e)}")

    def export_order_data(self):
        """导出订单数据"""
        try:
            # 获取选中行的数据
            selected_rows = set(item.row() for item in self.table_widget.selectedItems())
            if not selected_rows:
                QMessageBox.information(self, "提示", "请选择要导出的行")
                return
            
            # 获取选中行的实际数据
            selected_data = []
            for row in selected_rows:
                row_data = {}
                for col_idx, col_name in enumerate(self.original_column_names):
                    item = self.table_widget.item(row, col_idx)
                    if item:
                        row_data[col_name] = item.text().strip()
                    else:
                        row_data[col_name] = ""
                selected_data.append(row_data)
            
            # 显示导出对话框
            dialog = ExportOrderDialog(self, selected_data, self.table_name)
            if dialog.exec_() == QDialog.Accepted:
                export_format = dialog.get_export_format()
                filename = dialog.get_filename()
                include_headers = dialog.include_headers_checkbox.isChecked()
                
                # 执行导出
                self.perform_export(selected_data, export_format, filename, include_headers)
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"导出订单数据失败: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"导出订单数据失败: {str(e)}")

    def perform_export(self, data, export_format, filename, include_headers):
        """执行导出操作"""
        try:
            # 获取桌面路径作为默认导出目录
            import os
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(desktop_path):
                # 如果桌面路径不存在，使用当前工作目录
                desktop_path = os.getcwd()
            
            # 让用户选择导出目录
            from PyQt5.QtWidgets import QFileDialog
            selected_dir = QFileDialog.getExistingDirectory(
                self, 
                "选择导出目录", 
                desktop_path,
                QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
            )
            
            if not selected_dir:
                # 用户取消了选择
                return
            
            # 根据格式选择导出方法
            if export_format == "excel":
                self.export_to_excel(data, selected_dir, filename, include_headers)
            elif export_format == "csv":
                self.export_to_csv(data, selected_dir, filename, include_headers)
            else:  # txt
                self.export_to_txt(data, selected_dir, filename, include_headers)
            
            QMessageBox.information(self, "成功", f"数据已成功导出到: {selected_dir}")
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"执行导出失败: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def export_to_excel(self, data, export_dir, filename, include_headers):
        """导出为Excel格式"""
        try:
            import pandas as pd
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 构建完整文件路径
            file_path = os.path.join(export_dir, f"{filename}.xlsx")
            
            # 导出到Excel
            df.to_excel(file_path, index=not include_headers, engine='openpyxl')
            
        except ImportError:
            # 如果没有pandas，使用openpyxl直接导出
            self.export_to_excel_direct(data, export_dir, filename, include_headers)
        except Exception as e:
            raise Exception(f"导出Excel失败: {str(e)}")

    def export_to_excel_direct(self, data, export_dir, filename, include_headers):
        """使用openpyxl直接导出Excel（不依赖pandas）"""
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "订单数据"
            
            # 获取列名
            if data:
                columns = list(data[0].keys())
                
                # 写入表头
                if include_headers:
                    for col_idx, col_name in enumerate(columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    start_row = 2
                else:
                    start_row = 1
                
                # 写入数据
                for row_idx, row_data in enumerate(data, start_row):
                    for col_idx, col_name in enumerate(columns, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            
            # 保存文件
            file_path = os.path.join(export_dir, f"{filename}.xlsx")
            wb.save(file_path)
            
        except ImportError:
            raise Exception("导出Excel需要安装openpyxl库: pip install openpyxl")
        except Exception as e:
            raise Exception(f"导出Excel失败: {str(e)}")

    def export_to_csv(self, data, export_dir, filename, include_headers):
        """导出为CSV格式"""
        try:
            import csv
            
            # 构建完整文件路径
            file_path = os.path.join(export_dir, f"{filename}.csv")
            
            # 获取列名
            if not data:
                return
                
            columns = list(data[0].keys())
            
            # 写入CSV文件
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=columns)
                
                # 写入表头
                if include_headers:
                    writer.writeheader()
                
                # 写入数据
                writer.writerows(data)
                
        except Exception as e:
            raise Exception(f"导出CSV失败: {str(e)}")

    def export_to_txt(self, data, export_dir, filename, include_headers):
        """导出为TXT格式"""
        try:
            # 构建完整文件路径
            file_path = os.path.join(export_dir, f"{filename}.txt")
            
            # 获取列名
            if not data:
                return
                
            columns = list(data[0].keys())
            
            # 写入文本文件
            with open(file_path, 'w', encoding='utf-8') as txtfile:
                # 写入表头
                if include_headers:
                    header_line = "\t".join(columns)
                    txtfile.write(header_line + "\n")
                    txtfile.write("=" * len(header_line) + "\n\n")
                
                # 写入数据
                for row_data in data:
                    row_values = [str(row_data.get(col, "")) for col in columns]
                    line = "\t".join(row_values)
                    txtfile.write(line + "\n")
                
        except Exception as e:
            raise Exception(f"导出TXT失败: {str(e)}")

    def export_task_order_data(self, task_id, task_type):
        """导出任务相关的订单数据"""
        try:
            # 检查用户权限（只有爬虫权限才可使用）
            if not self.check_spider_permission():
                QMessageBox.warning(self, "权限不足", "只有爬虫权限才能使用此功能")
                return
            
            # 根据任务类型确定对应的表名
            table_name_map = {
                "帖子列表": "hupu_post_list",
                "帖子详情": "hupu_detail_list",
                "虎扑评分": "hupu_score_list"
            }
            table_name = table_name_map.get(task_type)
            if not table_name:
                QMessageBox.warning(self, "警告", f"不支持的任务类型: {task_type}")
                return
            
            # 直接显示导出对话框，不预先检查数据
            dialog = ExportOrderDialog(self, task_id=task_id, table_name=table_name)
            if dialog.exec_() == QDialog.Accepted:
                export_format = dialog.get_export_format()
                filename = dialog.get_filename()
                include_headers = dialog.include_headers_checkbox.isChecked()
                
                # 在用户确认导出后，再获取数据
                order_data = self.get_task_order_data(task_id, table_name)
                if not order_data:
                    QMessageBox.information(self, "提示", f"任务ID {task_id} 没有找到相关的订单数据")
                    return
                
                # 执行导出
                self.perform_export(order_data, export_format, filename, include_headers)
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"导出任务订单数据失败: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"导出任务订单数据失败: {str(e)}")

    def check_spider_permission(self):
        """检查用户是否有爬虫权限"""
        try:
            # 导入权限管理器
            from config.permission_manager import PermissionManager
            permission_manager = PermissionManager()
            
            # 检查当前用户是否有爬虫权限
            return permission_manager.has_permission("spider")
        except ImportError:
            # 如果没有权限管理器，默认允许
            return True
        except Exception as e:
            logger.error(f"检查权限失败: {str(e)}")
            # 出错时默认允许，避免影响正常使用
            return True

    def get_task_order_data(self, task_id, table_name):
        """获取任务相关的订单数据"""
        try:
            print(f"DEBUG: 获取任务订单数据 - task_id={task_id}, table_name={table_name}, db_path={self.db_path}")
            
            # 定义列别名映射（中文列名）
            column_aliases_map = {
                "hupu_post_list": {
                    "id": "ID",
                    "huputitle": "虎扑标题",
                    "hupu_zone": "虎扑分区",
                    "posturl": "帖子URL",
                    "replies": "回复数",
                    "tuijian_count": "推荐数",
                    "fatietime": "发帖时间",
                    "addtime": "添加时间",
                    "liangping_count": "亮评数",
                    "task_id": "任务ID"
                },
                "hupu_detail_list": {
                    "id": "ID",
                    "fabucontent": "发布内容",
                    "nickname": "昵称",
                    "replycontent": "回复内容",
                    "floor": "楼层",
                    "ipaddress": "IP地址",
                    "posttitle": "帖子标题",
                    "like_count": "点赞数",
                    "posturl": "帖子URL",
                    "replytime": "回复时间",
                    "addtime": "添加时间",
                    "task_id": "任务ID",
                    "reply_count": "回复数"
                },
                "hupu_score_list": {
                    "id": "ID",
                    "name": "名称",
                    "time": "时间",
                    "location": "位置",
                    "comment": "评论",
                    "reply_comment": "回复评论",
                    "like_count": "点赞数",
                    "score": "评分",
                    "score_title": "评分标题",
                    "addtime": "添加时间",
                    "task_id": "任务ID",
                    "scoreurl": "评分URL"
                }
            }
            
            # 获取当前表的列别名
            column_aliases = column_aliases_map.get(table_name, {})
            
            # 根据表名确定正确的数据库路径
            db_path = self.db_path
            if table_name in ["hupu_post_list", "hupu_detail_list", "hupu_score_list"]:
                # 虎扑相关表使用专门的数据库路径
                import os
                # 获取项目根目录
                current_dir = os.path.dirname(os.path.abspath(__file__))
                project_root = os.path.dirname(current_dir)
                hupu_db_config_path = os.path.join(project_root, "配置文件_系统配置", "hupu_db_config.json")
                
                print(f"DEBUG: 虎扑数据库配置文件路径 - {hupu_db_config_path}")
                
                # 读取虎扑数据库配置
                if os.path.exists(hupu_db_config_path):
                    try:
                        with open(hupu_db_config_path, "r", encoding="utf-8") as f:
                            hupu_db_config = json.load(f)
                            db_path = hupu_db_config.get("db_path", self.db_path)
                            # 确保路径是绝对路径
                            if not os.path.isabs(db_path):
                                db_path = os.path.join(project_root, db_path)
                            print(f"DEBUG: 使用虎扑数据库路径 - {db_path}")
                    except json.JSONDecodeError as e:
                        print(f"DEBUG: 虎扑数据库配置文件JSON解析失败 - {str(e)}")
                        # 使用默认路径
                        default_hupu_db_path = os.path.join(project_root, "配置文件_系统配置", "hupu.db")
                        if os.path.exists(default_hupu_db_path):
                            db_path = default_hupu_db_path
                            print(f"DEBUG: 使用默认虎扑数据库路径 - {db_path}")
                        else:
                            print(f"DEBUG: 默认虎扑数据库也不存在 - {default_hupu_db_path}")
                else:
                    print(f"DEBUG: 虎扑数据库配置文件不存在，创建默认配置")
                    # 创建默认的虎扑数据库配置文件
                    default_hupu_db_path = os.path.join(project_root, "配置文件_系统配置", "hupu.db")
                    default_config = {
                        "db_path": default_hupu_db_path,
                        "timeout": 30.0,
                        "check_same_thread": False,
                        "enable_foreign_keys": True,
                        "journal_mode": "WAL",
                        "cache_size": -20000,
                        "synchronous": "NORMAL",
                        "pool_config": {
                            "max_connections": 9999,
                            "min_connections": 1,
                            "connection_timeout": 30.0,
                            "idle_timeout": 300.0,
                            "pool_recycle": 3600,
                            "pool_pre_ping": True
                        },
                        "debug": False
                    }
                    
                    try:
                        with open(hupu_db_config_path, "w", encoding="utf-8") as f:
                            json.dump(default_config, f, indent=2, ensure_ascii=False)
                        print(f"DEBUG: 已创建虎扑数据库配置文件 - {hupu_db_config_path}")
                        db_path = default_hupu_db_path
                        print(f"DEBUG: 使用默认虎扑数据库路径 - {db_path}")
                    except Exception as config_error:
                        print(f"DEBUG: 创建虎扑数据库配置文件失败 - {str(config_error)}")
                        if os.path.exists(default_hupu_db_path):
                            db_path = default_hupu_db_path
                            print(f"DEBUG: 使用默认虎扑数据库路径 - {db_path}")
                        else:
                            print(f"DEBUG: 默认虎扑数据库也不存在 - {default_hupu_db_path}")
            
            # 使用数据库连接上下文管理器
            try:
                print(f"DEBUG: 尝试连接数据库 - {db_path}")
                # 直接连接SQLite数据库，不使用配置文件
                import sqlite3
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                
                # 构建查询SQL
                sql = f"SELECT * FROM {table_name} WHERE task_id = ? ORDER BY id"
                params = [task_id]
                
                print(f"DEBUG: 执行SQL - {sql}, params={params}")
                
                # 执行查询
                cursor.execute(sql, params)
                rows = cursor.fetchall()
                
                # 获取列名
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns_info = cursor.fetchall()
                columns = [col[1] for col in columns_info]
                
                print(f"DEBUG: 查询结果数量 - {len(rows)}")
                
                # 转换为字典列表
                order_data = []
                for row in rows:
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(columns):
                            col_name = columns[i]
                            # 使用列别名（中文）作为键
                            display_name = column_aliases.get(col_name, col_name)
                            row_dict[display_name] = str(value)
                    order_data.append(row_dict)
                
                conn.close()
                return order_data
            except Exception as db_error:
                print(f"DEBUG: 数据库操作失败 - {str(db_error)}")
                import traceback
                traceback.print_exc()
                return []
                
        except Exception as e:
            print(f"DEBUG: 获取任务订单数据异常 - {str(e)}")
            logger.error(f"获取任务订单数据失败: {str(e)}")
            return []

    def get_column_index_by_name(self, column_name):
        """根据列名获取列索引"""
        if column_name in self.original_column_names:
            return self.original_column_names.index(column_name)
        return -1

    def filter_order_by_taskid(self, order_tab, taskid):
        """在订单列表中根据任务ID筛选"""
        try:
            # 查找任务ID输入框
            if hasattr(order_tab, 'search_combo1') and hasattr(order_tab, 'search_input1'):
                # 设置筛选条件为任务ID
                index = order_tab.search_combo1.findText("任务ID")
                if index >= 0:
                    order_tab.search_combo1.setCurrentIndex(index)
                    order_tab.search_input1.setText(taskid)
                    order_tab.search_input2.clear()
                    order_tab.apply_search()
            else:
                QMessageBox.warning(self, "警告", "订单列表不支持筛选功能")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"筛选订单失败: {str(e)}")
    
    def show_header_context_menu(self, position):
        """显示表头右键菜单"""
        try:
            # 获取点击的列索引
            header = self.table_widget.horizontalHeader()
            column_index = header.logicalIndexAt(position)
            
            if column_index >= 0:
                menu = QMenu(self)
                
                # 添加自动调整列宽选项
                auto_resize_action = QAction("自动调整列宽", self)
                auto_resize_action.triggered.connect(lambda: self.auto_resize_column(column_index))
                menu.addAction(auto_resize_action)
                
                # 添加自动调整所有列宽选项
                auto_resize_all_action = QAction("自动调整所有列宽", self)
                auto_resize_all_action.triggered.connect(self.auto_resize_all_columns)
                menu.addAction(auto_resize_all_action)
                
                # 添加恢复默认列宽选项
                restore_default_action = QAction("恢复默认列宽", self)
                restore_default_action.triggered.connect(self.restore_default_column_widths)
                menu.addAction(restore_default_action)
                
                menu.exec_(header.mapToGlobal(position))
        except Exception as e:
            print(f"显示表头右键菜单时出错: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def auto_resize_column(self, column_index):
        """自动调整指定列的宽度以适应内容"""
        try:
            # 调整列宽
            self.table_widget.resizeColumnToContents(column_index)
            # 限制最小宽度，避免列太窄
            min_width = 80
            if self.table_widget.columnWidth(column_index) < min_width:
                self.table_widget.setColumnWidth(column_index, min_width)
                
            # 存储当前列宽设置，用于后续恢复
            if not hasattr(self, '_column_widths'):
                self._column_widths = {}
            self._column_widths[column_index] = self.table_widget.columnWidth(column_index)
            
            # 设置保持列宽标志
            self._keep_column_widths = True
        except Exception as e:
            print(f"自动调整列宽失败: {str(e)}")
    
    def auto_resize_all_columns(self):
        """自动调整所有列的宽度以适应内容"""
        try:
            # 调整所有列宽
            for column in range(self.table_widget.columnCount()):
                self.table_widget.resizeColumnToContents(column)
                # 限制最小宽度，避免列太窄
                min_width = 80
                if self.table_widget.columnWidth(column) < min_width:
                    self.table_widget.setColumnWidth(column, min_width)
                    
            # 存储当前列宽设置，用于后续恢复
            if not hasattr(self, '_column_widths'):
                self._column_widths = {}
            for column in range(self.table_widget.columnCount()):
                self._column_widths[column] = self.table_widget.columnWidth(column)
            
            # 设置保持列宽标志
            self._keep_column_widths = True
        except Exception as e:
            print(f"自动调整所有列宽失败: {str(e)}")
    
    def resizeEvent(self, event):
        """重写resizeEvent，保持自定义列宽"""
        super().resizeEvent(event)
        
        # 只有在设置了保持列宽标志时才恢复列宽
        if hasattr(self, '_keep_column_widths') and self._keep_column_widths:
            if hasattr(self, '_column_widths') and self._column_widths:
                # 禁用自动拉伸
                self.table_widget.horizontalHeader().setStretchLastSection(False)
                
                # 恢复列宽
                for column, width in self._column_widths.items():
                    self.table_widget.setColumnWidth(column, width)
                    
                # 重新启用自动拉伸，但只对最后一列生效
                self.table_widget.horizontalHeader().setStretchLastSection(True)
    
    def restore_default_column_widths(self):
        """恢复默认列宽"""
        try:
            # 清除自定义列宽设置
            if hasattr(self, '_column_widths'):
                delattr(self, '_column_widths')
            
            # 设置不保持列宽标志
            self._keep_column_widths = False
            
            # 重置为默认状态，让表格自动管理列宽
            self.table_widget.horizontalHeader().setStretchLastSection(True)
            
            # 可选：自动调整所有列宽到内容
            for column in range(self.table_widget.columnCount()):
                self.table_widget.resizeColumnToContents(column)
                # 限制最小宽度，避免列太窄
                min_width = 80
                if self.table_widget.columnWidth(column) < min_width:
                    self.table_widget.setColumnWidth(column, min_width)
        except Exception as e:
            print(f"恢复默认列宽失败: {str(e)}")

    def open_related_data_by_task_type(self):
        """根据任务类型打开相关数据页面 - 跳转到对应标签页并设置任务ID搜索"""
        try:
            # 获取选中行的任务ID和任务类型
            selected_ids = self.get_selected_task_ids()
            if not selected_ids:
                QMessageBox.information(self, "提示", "请选择要打开的任务")
                return
            
            task_id = selected_ids[0]  # 使用第一个选中的任务ID
            
            # 获取任务类型
            task_type = self.get_task_type_by_id(task_id)
            if not task_type:
                QMessageBox.warning(self, "警告", "无法确定任务类型")
                return
            
            # 根据任务类型决定打开哪个页面
            target_table = None
            if task_type == "帖子列表":
                target_table = "hupu_post_list"
            elif task_type == "帖子详情":
                target_table = "hupu_detail_list"
            elif task_type == "虎扑评分":
                target_table = "hupu_score_list"
            else:
                QMessageBox.information(self, "提示", f"任务类型 '{task_type}' 没有对应的数据页面")
                return
            
            # 获取父级DbTableViewer实例
            parent_viewer = self.get_parent_viewer()
            if not parent_viewer:
                QMessageBox.warning(self, "警告", "无法访问标签页控件")
                return
            
            # 查找目标标签页
            target_tab = None
            for i in range(parent_viewer.tab_widget.count()):
                tab_widget = parent_viewer.tab_widget.widget(i)
                if hasattr(tab_widget, 'table_name') and tab_widget.table_name == target_table:
                    target_tab = tab_widget
                    break
            
            if not target_tab:
                QMessageBox.warning(self, "警告", f"找不到目标标签页: {target_table}")
                return
            
            # 切换到目标标签页
            parent_viewer.tab_widget.setCurrentWidget(target_tab)
            
            # 设置搜索条件并刷新数据 - 参考SqlList.py的loadOrderData实现
            if hasattr(target_tab, 'search_combo1') and hasattr(target_tab, 'search_input1'):
                # 设置筛选条件为任务ID
                index = target_tab.search_combo1.findText("任务ID")
                if index >= 0:
                    target_tab.search_combo1.setCurrentIndex(index)
                    target_tab.search_input1.setText(task_id)
                    target_tab.search_input2.clear()
                    
                    # 直接刷新数据，不使用apply_search，避免分页问题
                    if hasattr(target_tab, 'load_data'):
                        target_tab.load_data()
                    elif hasattr(target_tab, 'refresh_data'):
                        target_tab.refresh_data()
            else:
                QMessageBox.warning(self, "警告", "目标标签页不支持搜索功能")
            
            QMessageBox.information(self, "成功", f"已跳转到{task_type}数据页面，并设置任务ID为 {task_id}")
            
        except Exception as e:
            # 获取详细的错误信息
            import traceback
            error_details = traceback.format_exc()
            error_message = f"打开相关数据失败: {str(e)}"
            
            # 记录详细错误信息
            logger.error(f"任务管理页面打开相关数据错误: {error_message}\n{error_details}")
            
            # 显示错误对话框
            QMessageBox.critical(self, "错误", f"{error_message}\n\n详细信息已记录到日志文件")
    
    def get_task_type_by_id(self, task_id):
        """根据任务ID获取任务类型"""
        try:
            # 使用全局数据库连接管理器获取task表的连接
            from config.common_config import db_manager
            db = db_manager.get_connection("task")
            
            if not db:
                logger.error("无法获取task表的数据库连接")
                return None
            
            # 查询数据库获取任务名称和函数名称
            # 先尝试通过task_id查询，如果失败则通过id查询
            query = "SELECT task_name, func_name FROM task WHERE task_id = ?"
            result = db.execute_sql(query, params=(task_id,), fetch="fetch_one")
            
            if not result:
                # 如果通过task_id没找到，尝试通过id查询
                query = "SELECT task_name, func_name FROM task WHERE id = ?"
                result = db.execute_sql(query, params=(task_id,), fetch="fetch_one")
            
            if result:
                # 优先使用task_name，如果为空则使用func_name
                task_name = result.get("task_name", "").strip()
                func_name = result.get("func_name", "").strip()
                
                # 用于显示的名称（优先task_name）
                display_name = task_name if task_name else func_name
                
                # 根据名称判断任务类型
                if "虎扑帖子列表" in display_name:
                    return "帖子列表"
                elif "虎扑帖子详情" in display_name:
                    return "帖子详情"
                elif "虎扑评分" in display_name:
                    return "虎扑评分"
                else:
                    return display_name  # 如果无法匹配，返回显示名称
            return None
        except Exception as e:
            # 获取详细的错误信息
            import traceback
            error_details = traceback.format_exc()
            error_message = f"获取任务类型失败: {str(e)}"
            
            # 记录详细错误信息
            logger.error(f"任务管理页面获取任务类型错误: {error_message}\n{error_details}")
            
            # 返回None而不是抛出异常
            return None
    
    def search_in_related_tables(self, target_table_name):
        """在相关表中搜索任务ID"""
        try:
            # 获取选中行的任务ID
            selected_ids = self.get_selected_task_ids()
            if not selected_ids:
                QMessageBox.information(self, "提示", "请选择要搜索的任务")
                return
            
            task_id = selected_ids[0]  # 使用第一个选中的任务ID
            
            # 获取父级DbTableViewer实例 - 使用更可靠的方式
            parent_viewer = self.get_parent_viewer()
            if not parent_viewer:
                QMessageBox.warning(self, "警告", "无法访问标签页控件")
                return
            
            # 查找目标标签页
            target_tab = None
            for i in range(parent_viewer.tab_widget.count()):
                tab_widget = parent_viewer.tab_widget.widget(i)
                if hasattr(tab_widget, 'table_name') and tab_widget.table_name == target_table_name:
                    target_tab = tab_widget
                    parent_viewer.tab_widget.setCurrentIndex(i)
                    break
            
            if not target_tab:
                QMessageBox.warning(self, "警告", f"未找到目标表: {target_table_name}")
                return
            
            # 等待标签页加载完成，然后设置搜索条件
            QTimer.singleShot(500, lambda: self.set_search_condition(target_tab, task_id))
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"搜索失败: {str(e)}")
    
    def get_parent_viewer(self):
        """获取父级DbTableViewer实例 - 使用更可靠的方式"""
        try:
            # 方法1：尝试通过parent属性逐级查找
            current = self
            for _ in range(5):  # 最多向上查找5级
                if hasattr(current, 'tab_widget') and hasattr(current, 'tab_configs'):
                    return current  # 找到DbTableViewer实例
                current = current.parent()
                if current is None:
                    break
            
            # 方法2：如果方法1失败，尝试通过全局窗口查找
            from PyQt5.QtWidgets import QApplication
            for widget in QApplication.topLevelWidgets():
                if isinstance(widget, DbTableViewer):
                    # 检查这个DbTableViewer是否包含当前标签页
                    for i in range(widget.tab_widget.count()):
                        if widget.tab_widget.widget(i) == self:
                            return widget
            
            return None
        except Exception as e:
            print(f"获取父级DbTableViewer失败: {str(e)}")
            return None

    def set_search_condition(self, target_tab, task_id):
        """在目标标签页中设置搜索条件 - 安全地设置任务ID搜索"""
        try:
            # 检查目标标签页是否有搜索功能
            if not hasattr(target_tab, 'search_combo1') or not hasattr(target_tab, 'search_input1'):
                # 如果没有搜索功能，尝试直接设置过滤条件
                if hasattr(target_tab, 'set_filter'):
                    target_tab.set_filter('task_id', task_id)
                elif hasattr(target_tab, 'apply_search'):
                    # 尝试调用apply_search方法
                    target_tab.search_term = task_id
                    target_tab.apply_search()
                else:
                    print(f"[警告] 目标标签页不支持搜索功能: {target_tab}")
                return
            
            # 搜索任务ID
            index = target_tab.search_combo1.findText("任务ID")
            if index >= 0:
                target_tab.search_combo1.setCurrentIndex(index)
                target_tab.search_input1.setText(task_id)
                target_tab.search_input2.clear()
                
                # 尝试应用搜索
                if hasattr(target_tab, 'apply_search'):
                    target_tab.apply_search()
                elif hasattr(target_tab, 'search'):
                    target_tab.search()
                else:
                    print(f"[警告] 目标标签页没有apply_search或search方法")
            else:
                # 如果找不到任务ID选项，尝试其他方式
                print(f"[警告] 找不到任务ID搜索选项，尝试其他方式")
                
                # 尝试直接设置搜索条件
                if hasattr(target_tab, 'search_input1'):
                    target_tab.search_input1.setText(task_id)
                    
                    # 尝试应用搜索
                    if hasattr(target_tab, 'apply_search'):
                        target_tab.apply_search()
                    elif hasattr(target_tab, 'search'):
                        target_tab.search()
                    
        except Exception as e:
            print(f"[错误] 设置搜索条件失败: {str(e)}")
            # 不显示错误对话框，避免中断用户操作

    def get_selected_task_ids(self) -> List[str]:
        """获取选中行的task_id列表（用于task表）"""
        selected_rows = set(item.row() for item in self.table_widget.selectedItems())
        task_ids = []

        # 首先尝试通过列名获取task_id
        task_id_col_index = -1
        for col in range(self.table_widget.columnCount()):
            header_text = self.table_widget.horizontalHeaderItem(col).text()
            if header_text == "任务ID" or header_text == "task_id":
                task_id_col_index = col
                break

        for row in selected_rows:
            task_id = ""
            
            # 如果找到了task_id列，直接从该列获取
            if task_id_col_index >= 0:
                item = self.table_widget.item(row, task_id_col_index)
                if item:
                    task_id = item.text().strip()
            else:
                # 如果没有找到task_id列，尝试从每列中查找包含task_id格式的内容
                for col in range(self.table_widget.columnCount()):
                    item = self.table_widget.item(row, col)
                    if item and item.text():
                        text = item.text().strip()
                        # 检查是否符合task_id的格式（通常是32位的十六进制字符串）
                        if len(text) == 32 and all(c in "0123456789abcdefABCDEF" for c in text):
                            task_id = text
                            break
            
            if task_id:
                task_ids.append(task_id)

        return task_ids

    # 分页相关方法
    def update_pagination_info(self):
        """更新分页信息显示"""
        # 使用实际加载的数据条数
        actual_current_page_count = len(self.original_data) if hasattr(self, 'original_data') and self.original_data else 0
        
        # 根据是否处于搜索状态显示不同的格式
        if hasattr(self, 'is_searching') and self.is_searching:
            # 搜索状态：显示找到x条，总共y条，当前页z条
            search_result_count = getattr(self, 'search_result_count', 0)
            total_record_count = getattr(self, 'total_record_count', 0)
            self.current_data_info_label.setText(f"找到{search_result_count}条，总共{total_record_count}条，当前页{actual_current_page_count}条")
        else:
            # 非搜索状态：显示总共y条，当前页z条
            total_record_count = getattr(self, 'total_record_count', 0)
            self.current_data_info_label.setText(f"总共{total_record_count}条，当前页{actual_current_page_count}条")
        
        # 更新右侧页码信息标签
        self.page_info_label.setText(f"共{self.total_pages}页，当前第{self.current_page}页")
        
        # 更新页码输入框 - 使用QLineEdit的setText方法
        self.page_input.setText(str(self.current_page))
        
        # 更新按钮状态
        self.prev_page_btn.setEnabled(self.current_page > 1)
        self.next_page_btn.setEnabled(self.current_page < self.total_pages)
        self.go_to_page_btn.setEnabled(self.total_pages > 1)

    def go_to_page(self):
        """前往指定页码"""
        text = self.page_input.text().strip()
        if not text:
            return
        
        try:
            page = int(text)
            if page < 1 or page > self.total_pages:
                QMessageBox.warning(self, "警告", f"页码超出范围，请输入1-{self.total_pages}之间的数字")
                self.page_input.setText(str(self.current_page))
                return
            
            self.current_page = page
            self.load_data()
        except ValueError:
            QMessageBox.warning(self, "警告", "请输入有效的页码数字")
            self.page_input.setText(str(self.current_page))

    def go_to_prev_page(self):
        """前往上一页"""
        if self.current_page > 1:
            self.current_page -= 1
            self.load_data()

    def go_to_next_page(self):
        """前往下一页"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_data()
    
    def on_page_size_changed(self, text):
        """分页大小变化时处理"""
        if not text:
            return
        
        try:
            new_page_size = int(text)
            if new_page_size < 1:
                # 如果分页大小小于1，恢复默认值
                self.page_size_combo.setCurrentText(str(self.page_size))
                return
            
            # 更新分页大小
            self.page_size = new_page_size
            
            # 重置到第一页
            self.current_page = 1
            
            # 重新加载数据
            self.load_data()
        except ValueError:
            # 如果输入的不是数字，恢复当前值
            self.page_size_combo.setCurrentText(str(self.page_size))
    
    def copy_cell_content(self, item):
        """复制单元格内容到剪贴板"""
        if item and isinstance(item, CustomTableWidgetItem):
            # 获取原始值，不是显示文本
            content = item.sort_key
            # 如果是字典或列表，格式化为JSON
            if isinstance(content, (dict, list)):
                content_str = json.dumps(content, ensure_ascii=False, indent=2)
            else:
                content_str = str(content)
            
            # 复制到剪贴板
            clipboard = QApplication.clipboard()
            clipboard.setText(content_str)
            
            # 显示提示
            self.show_copy_message("单元格内容已复制")
    
    def copy_selected_rows(self):
        """复制选中行数据到剪贴板"""
        # 获取选中的行（使用更可靠的方法）
        selected_indexes = self.table_widget.selectionModel().selectedRows()
        selected_rows = [index.row() for index in selected_indexes]
        
        if not selected_rows:
            self.show_copy_message("请先选择要复制的行")
            return
        
        # 获取表头
        headers = [self.column_aliases.get(col, col) for col in self.original_column_names]
        
        # 收集所有选中行的数据
        rows_data = []
        for row in selected_rows:
            row_data = []
            for col_idx, col_name in enumerate(self.original_column_names):
                item = self.table_widget.item(row, col_idx)
                if item and isinstance(item, CustomTableWidgetItem):
                    # 获取原始值
                    value = item.sort_key
                    # 格式化值
                    if isinstance(value, (dict, list)):
                        value_str = json.dumps(value, ensure_ascii=False)
                    elif value is None:
                        value_str = ""
                    else:
                        value_str = str(value)
                    row_data.append(value_str)
                else:
                    row_data.append("")
            rows_data.append(row_data)
        
        # 格式化为表格文本
        if len(rows_data) == 1:
            # 单行，直接复制
            result_text = "\t".join(rows_data[0])
        else:
            # 多行，添加表头
            result_text = "\t".join(headers) + "\n"
            for row_data in rows_data:
                result_text += "\t".join(row_data) + "\n"
        
        # 复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(result_text)
        
        # 显示提示
        self.show_copy_message(f"已复制 {len(rows_data)} 行数据")
    
    def show_copy_message(self, message):
        """显示复制提示消息"""
        # 获取主窗口
        main_window = self.parent()
        while main_window and main_window.parent():
            main_window = main_window.parent()
        
        # 查找状态栏
        status_bar = None
        if main_window and hasattr(main_window, 'statusBar'):
            status_bar = main_window.statusBar()
        
        if status_bar:
            # 保存原消息
            original_message = status_bar.currentMessage()
            
            # 显示复制消息
            status_bar.showMessage(message, 2000)  # 显示2秒
            
            # 2秒后恢复原消息
            from PyQt5.QtCore import QTimer
            QTimer.singleShot(2000, lambda: status_bar.showMessage(original_message))
        else:
            # 如果没有状态栏，使用控制台输出
            print(message)
    
    def open_url_in_browser(self, url):
        """在浏览器中打开URL"""
        if not url:
            self.show_copy_message("没有可打开的URL")
            return False
            
        # 使用lite_modules中的核心函数
        from lite_modules.web_utils import open_url_in_browser_core
        success = open_url_in_browser_core(url)
        
        if success:
            self.show_copy_message(f"已打开URL: {url}")
        else:
            self.show_copy_message(f"打开URL失败: {url}")
            
        return success
    
    def open_hupu_score_homepage(self):
        """打开虎扑评分首页"""
        try:
            url = "https://bbsactivity.hupu.com/pc-viewer/index.html"
            from lite_modules.web_utils import open_url_in_browser_core
            success = open_url_in_browser_core(url)
            if success:
                self.show_copy_message("已打开虎扑评分首页")
            else:
                self.show_copy_message("打开虎扑评分首页失败")
        except Exception as e:
            self.show_copy_message(f"打开虎扑评分首页失败: {str(e)}")
    
    def open_hupu_post_search_homepage(self):
        """打开帖子搜索首页"""
        try:
            url = "https://bbs.hupu.com/search?q=%E5%B0%8F%E9%BB%91%E5%AD%90&topicId=&sortby=reply&page=1"
            from lite_modules.web_utils import open_url_in_browser_core
            success = open_url_in_browser_core(url)
            if success:
                self.show_copy_message("已打开帖子搜索首页")
            else:
                self.show_copy_message("打开帖子搜索首页失败")
        except Exception as e:
            self.show_copy_message(f"打开帖子搜索首页失败: {str(e)}")
    
    def get_url_from_cell(self, row, col):
        """从指定单元格获取URL"""
        item = self.table_widget.item(row, col)
        if item and isinstance(item, CustomTableWidgetItem):
            url = item.sort_key
            if isinstance(url, str) and url.strip():
                # 清理URL，去除反引号等特殊字符
                url = url.strip().strip('`').strip('"').strip("'")
                return url
        return None
    
    def add_url_menu_items(self, menu, position):
        """添加URL相关的菜单项"""
        try:
            # 获取点击位置的行
            item = self.table_widget.itemAt(position)
            if not item:
                return
                
            row = item.row()
            
            # 根据表格名称确定URL列和标签
            table_name = self.table_name.lower()
            
            # 帖子列表和帖子详情表：帖子URL
            if 'post' in table_name:
                url_columns = ['posturl']  # 使用原始列名
                url_label = '打开帖子链接'
            # 虎扑评分表：评分URL
            elif 'score' in table_name:
                url_columns = ['scoreurl']  # 使用原始列名
                url_label = '打开评分链接'
            # 其他表：通用URL
            else:
                url_columns = ['posturl', 'scoreurl', 'url']
                url_label = '打开链接'
            
            # 查找该行中的URL列
            found_urls = []
            if hasattr(self, 'original_column_names') and self.original_column_names:
                for i, name in enumerate(self.original_column_names):
                    if name.lower() in url_columns:
                        url = self.get_url_from_cell(row, i)
                        if url:
                            # 获取列的显示别名
                            display_name = self.column_aliases.get(name, name)
                            found_urls.append((name, display_name, url))
            
            # 为每个找到的URL添加菜单项
            for col_name, display_name, url in found_urls:
                # 使用显示别名而不是列名
                action_text = f"{url_label}: {display_name}" if 'url' not in url_label.lower() else url_label
                open_url_action = QAction(action_text, self)
                
                # 使用functools.partial来避免lambda闭包问题
                from functools import partial
                open_url_action.triggered.connect(partial(self.open_url_in_browser, url))
                menu.addAction(open_url_action)
            
            # 如果找到URL，添加分隔线
            if found_urls:
                menu.addSeparator()
        except Exception as e:
            # 如果出现异常，打印错误信息但不影响其他菜单项
            print(f"添加URL菜单项时出错: {str(e)}")
            import traceback
            traceback.print_exc()


class DbTableViewer(QMainWindow):
    """主窗口，包含分页选项卡"""

    def __init__(self, db_path: str, tab_configs: List[Dict]):
        super().__init__()
        self.db_path = db_path
        self.tab_configs = tab_configs
        self.setWindowIcon(QIcon("gui/img/favicon.ico"))
        # 初始化UI
        self.init_main_ui()

    def init_main_ui(self):
        """初始化主窗口UI - 保持原有样式"""
        self.setWindowTitle("Ikun联盟 - 数据库")
        self.resize(1400, 800)
        self.center_window()

        # 中央组件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 分页选项卡
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)

        # 创建所有标签页
        self.create_tabs()

    def center_window(self):
        """窗口居中 - 保持原有逻辑"""
        screen_rect = QApplication.desktop().screenGeometry()
        window_rect = self.frameGeometry()
        window_rect.moveCenter(screen_rect.center())
        self.move(window_rect.topLeft())

    def create_tabs(self):
        """创建所有配置的标签页 - 支持每个标签页使用不同的数据库"""
        for tab_config in self.tab_configs:
            # 获取数据库路径，如果没有指定则使用默认路径
            tab_db_path = tab_config.get("db_path")
            if tab_db_path is None:
                tab_db_path = self.db_path

            # 创建标签页内容
            tab_content = TableTabWidget(
                parent=self,
                db_path=tab_db_path,
                table_name=tab_config["table_name"],
                columns_to_display=tab_config["columns_to_display"],
                column_aliases=tab_config["column_aliases"],
                column_width_config=tab_config["column_width_config"],
                context_menu_actions=tab_config["context_menu_actions"],
                tab_name=tab_config["tab_name"]
            )

            # 添加到分页控件
            self.tab_widget.addTab(tab_content, tab_config["tab_name"])

    # ------------------------------ 通用操作方法 - 完全使用execute_sql重构 ------------------------------
    def modify_field(self, table_name: str, field_name: str, field_display: str, selected_ids: List[int]):
        """通用修改字段方法 - 使用execute_sql重构，支持多数据库"""
        if not selected_ids:
            QMessageBox.information(self, "提示", "请选择要修改的行")
            return

        new_value, ok = QInputDialog.getText(self, f"修改{field_display}", f"新的{field_display}:")
        if ok and new_value:
            db = None
            try:
                # 获取当前活动标签页的数据库路径
                current_tab = self.tab_widget.currentWidget()
                db_path = current_tab.db_path
                
                db = SQLiteDB(db_path)

                # 构建参数化SQL，防止注入
                placeholders = ", ".join("?" * len(selected_ids))
                sql = f"UPDATE {table_name} SET {field_name} = ? WHERE id IN ({placeholders})"
                params = [new_value] + selected_ids

                # 使用execute_sql执行更新
                affected = db.execute_sql(sql, params=params, fetch="none")

                QMessageBox.information(self, "成功", f"修改了 {affected} 行")
                # 刷新当前标签页
                current_tab = self.tab_widget.currentWidget()
                current_tab.load_data()

            except Exception as e:
                QMessageBox.critical(self, "错误", f"修改失败: {str(e)}")
            finally:
                if db is not None:
                    try:
                        db.close()
                    except:
                        # 忽略关闭连接时的错误
                        pass

    def clear_shop_upload_pic_spu_record(self, selected_ids: List[int] = None):
        """
        清空SPU记录（先查真实UID再筛选，支持「按选中店铺精准清空」+「全局清空」）
        :param selected_ids: shops表选中行的id列表（数字），传None/空则全局清空（执行DELETE FROM record where 1）
        核心步骤：1. 用selected_ids查shops表获取真实uid；2. 用uid筛选record表删除（兼容数字/字符串uid）
        """
        # 1. 区分操作类型，初始化变量
        is_global = False
        where_cond = ""
        title = ""
        msg = ""
        real_uids = []  # 存储从shops表查询到的真实uid

        if selected_ids and len(selected_ids) > 0:
            # 按选中店铺清空：先查真实uid，再构建筛选条件
            title = "确认清空选中店铺SPU"
            msg = f"确定要清空【共{len(selected_ids)}个店铺】的SPU记录吗？此操作不可恢复！"

            try:
                # 第一步：通过shops表id查询真实uid（核心！参数化SQL防注入）
                db = SQLiteDB(self.db_path)
                placeholders = ", ".join("?" * len(selected_ids))
                query_uid_sql = f"SELECT uid FROM shops WHERE id IN ({placeholders})"
                # 执行查询，获取店铺真实uid列表
                uid_result = db.execute_sql(query_uid_sql, params=selected_ids, fetch="fetch")
                
                # 提取真实uid，去重（避免重复删除）
                real_uids = [row.get("uid") for row in uid_result if row.get("uid") is not None]
                if not real_uids:
                    QMessageBox.information(self, "提示", "未查询到选中店铺对应的UID，无需清空！")
                    return

                # 第二步：构建兼容数字/字符串uid的筛选条件
                or_conditions = []
                for uid in real_uids:
                    if isinstance(uid, (int, float)):
                        # 数字类型uid：直接匹配 + 字符串化匹配（兼容record表uid为字符串的情况）
                        or_conditions.append(f"uid = {uid} OR uid = '{uid}'")
                    else:
                        # 字符串类型uid：单引号包裹，避免SQL语法错误
                        or_conditions.append(f"uid = '{uid}'")
                where_cond = "WHERE " + " OR ".join(or_conditions)

            except Exception as e:
                QMessageBox.critical(self, "查询失败", f"查询店铺UID出错：{str(e)}")
                return
            finally:
                if db is not None:
                    try:
                        db.close()
                    except:
                        # 忽略关闭连接时的错误
                        pass
        else:
            # 全局清空：完全复用原始逻辑，无筛选条件
            is_global = True
            title = "确认清空所有SPU"
            msg = "确定要清空record表中所有SPU记录吗？此操作不可恢复！"
            where_cond = "where 1"

        # 2. 确认弹窗（防误操作，默认选中取消）
        reply = QMessageBox.question(
            self, title, msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No  # 默认聚焦取消，防止回车误点
        )
        if reply != QMessageBox.Yes:
            return

        # 3. 执行删除操作
        db = None
        try:
            db = SQLiteDB(self.db_path)
            # 拼接最终删除SQL，核心复用原始清理逻辑
            delete_sql = f"DELETE FROM record {where_cond}"
            print("执行删除SQL:", delete_sql)  # 调试用，可删除
            # 执行删除，返回受影响行数
            affected_rows = db.execute_sql(delete_sql, fetch="none")

            # 4. 操作结果反馈（区分场景）
            if affected_rows > 0:
                if is_global:
                    tip = "所有"
                else:
                    tip = f"选中的{len(selected_ids)}个店铺（匹配到{len(real_uids)}个有效UID）"
                QMessageBox.information(self, "操作成功", f"已成功清空{tip}SPU记录，共删除 {affected_rows} 条数据！")
            else:
                QMessageBox.information(self, "操作提示", "未找到匹配的SPU记录，无需清空！")

            # 5. 刷新当前标签页，数据实时更新
            current_tab = self.tab_widget.currentWidget()
            current_tab.load_data()

        except Exception as e:
            QMessageBox.critical(self, "操作失败", f"清空SPU记录出错：{str(e)}")
        finally:
            if db is not None:
                try:
                    db.close()
                except:
                    # 忽略关闭连接时的错误
                    pass

    def clear_auth(self, table_name: str, selected_ids: List[int]):
        """通用清空认证信息方法 - 使用execute_sql重构，支持多数据库"""
        if not selected_ids:
            return

        reply = QMessageBox.question(self, "确认", "确定清空认证信息？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            db = None
            try:
                # 获取当前活动标签页的数据库路径
                current_tab = self.tab_widget.currentWidget()
                db_path = current_tab.db_path
                
                db = SQLiteDB(db_path)

                # 构建参数化SQL
                placeholders = ", ".join("?" * len(selected_ids))
                sql = f"""
                    UPDATE {table_name} 
                    SET headers = NULL, cookies = NULL, connect_status = '未连接' 
                    WHERE id IN ({placeholders})
                """

                # 使用execute_sql执行更新
                affected = db.execute_sql(sql, params=selected_ids, fetch="none")

                QMessageBox.information(self, "成功", f"清空了 {affected} 行认证信息")
                self.tab_widget.currentWidget().load_data()

            except Exception as e:
                QMessageBox.critical(self, "错误", f"操作失败: {str(e)}")
            finally:
                if db is not None:
                    try:
                        db.close()
                    except:
                        # 忽略关闭连接时的错误
                        pass

    def delete_rows(self, table_name: str, selected_ids: List[int]):
        """通用删除行方法 - 使用execute_sql重构，支持多数据库"""
        if not selected_ids:
            return

        # 根据当前标签页显示不同的确认提示
        current_tab = self.tab_widget.currentWidget()
        if current_tab.table_name == "task":
            # 任务管理表的特殊提示
            confirm_msg = "该方法不会停止任务线程，确定删除？"
        else:
            # 其他表的通用提示
            confirm_msg = "确认删除吗？操作无法恢复！"
            
        reply = QMessageBox.question(self, "确认删除", confirm_msg, QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            db = None
            try:
                # 获取当前活动标签页的数据库路径
                db_path = current_tab.db_path
                
                db = SQLiteDB(db_path)

                # 构建参数化SQL
                placeholders = ", ".join("?" * len(selected_ids))
                sql = f"DELETE FROM {table_name} WHERE id IN ({placeholders})"

                # 使用execute_sql执行删除
                affected = db.execute_sql(sql, params=selected_ids, fetch="none")

                QMessageBox.information(self, "成功", f"删除了 {affected} 行")
                self.tab_widget.currentWidget().load_data()

            except Exception as e:
                QMessageBox.critical(self, "错误", f"删除失败: {str(e)}")
            finally:
                if db is not None:
                    try:
                        db.close()
                    except:
                        # 忽略关闭连接时的错误
                        pass

    def export_selected_rows(self, table_name: str, selected_ids: List[int]):
        """导出选中行"""
        if not selected_ids:
            QMessageBox.information(self, "提示", "请选择要导出的行")
            return
        
        try:
            # 获取当前活动标签页
            current_tab = self.tab_widget.currentWidget()
            
            # 获取选中行的数据
            selected_data = []
            for row_id in selected_ids:
                # 在表格中查找对应行的数据
                for row_idx in range(current_tab.table_widget.rowCount()):
                    item = current_tab.table_widget.item(row_idx, 0)
                    if item and int(item.text()) == row_id:
                        row_data = {}
                        for col_idx, col_name in enumerate(current_tab.original_column_names):
                            item = current_tab.table_widget.item(row_idx, col_idx)
                            if item:
                                # 使用列别名（中文）作为键
                                display_name = current_tab.column_aliases.get(col_name, col_name)
                                row_data[display_name] = item.text()
                            else:
                                display_name = current_tab.column_aliases.get(col_name, col_name)
                                row_data[display_name] = ""
                        selected_data.append(row_data)
                        break
            
            if not selected_data:
                QMessageBox.warning(self, "警告", "未找到选中的数据")
                return
            
            # 显示导出对话框
            dialog = ExportOrderDialog(self, table_name=table_name, data=selected_data)
            if dialog.exec_() == QDialog.Accepted:
                export_format = dialog.get_export_format()
                filename = dialog.get_filename()
                include_headers = dialog.include_headers_checkbox.isChecked()
                
                # 执行导出 - 调用当前标签页的perform_export方法
                current_tab.perform_export(selected_data, export_format, filename, include_headers)
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"导出选中行失败: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def modify_task_status(self, selected_ids: List[int]):
        """修改任务状态 - 使用数据库索引id（参考multiThreading_log_manager.py写法），支持多数据库"""
        if not selected_ids:
            QMessageBox.information(self, "提示", "请选择要修改的任务")
            return

        # 创建对话框
        dialog = TaskStatusDialog(self, [str(id) for id in selected_ids])  # 对话框需要字符串列表
        if dialog.exec_() == QDialog.Accepted:
            new_status = dialog.get_selected_status()
            if new_status:
                try:
                    # 使用全局数据库连接管理器获取task表的连接
                    from config.common_config import db_manager
                    db = db_manager.get_connection("task")
                    
                    if not db:
                        QMessageBox.critical(self, "错误", "无法获取数据库连接")
                        return
                    
                    # 参考multiThreading_log_manager.py的update_task_field方法写法
                    # 构建参数化SQL，防止SQL注入，使用数据库索引id
                    placeholders = ", ".join("?" * len(selected_ids))
                    update_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    sql = f"""
                        UPDATE task 
                        SET status = ?,
                            msg = ?,
                            remarks = ?,
                            update_time = datetime('now', '+8 hours')
                        WHERE id IN ({placeholders})
                    """
                    msg = f"状态已修改为: {new_status}"
                    remarks = f"通过数据库管理界面手动修改状态 | 修改时间: {update_time}"
                    params = [new_status, msg, remarks] + selected_ids
                    
                    # 使用execute_sql执行更新（fetch="none"表示不获取结果，commit=True默认会自动提交）
                    affected_rows = db.execute_sql(sql, params=params, fetch="none")

                    # 显示结果（affected_rows是受影响的行数）
                    if affected_rows and affected_rows > 0:
                        QMessageBox.information(self, "成功", f"成功修改 {affected_rows} 个任务的状态为: {new_status}")
                    else:
                        QMessageBox.warning(self, "警告", f"未更新任何任务，请检查选中的行是否正确")
                    
                    # 刷新当前标签页
                    current_tab = self.tab_widget.currentWidget()
                    current_tab.load_data()

                except Exception as e:
                    import traceback
                    error_msg = f"修改状态失败: {str(e)}"
                    QMessageBox.critical(self, "错误", error_msg)
                    # 可选：打印详细错误信息用于调试
                    print(f"修改任务状态错误详情:\n{traceback.format_exc()}")

    def delete_task_rows(self, selected_ids: List[int]):
        """删除task表的行（使用数据库索引id）"""
        if not selected_ids:
            return

        reply = QMessageBox.question(self, "确认删除", "该方法不会停止任务线程，确定删除？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            db = None
            try:
                db = SQLiteDB(self.db_path)

                # 构建参数化SQL（使用数据库索引id）
                placeholders = ", ".join("?" * len(selected_ids))
                sql = f"DELETE FROM task WHERE id IN ({placeholders})"

                # 使用execute_sql执行删除
                affected = db.execute_sql(sql, params=selected_ids, fetch="none")

                QMessageBox.information(self, "成功", f"删除了 {affected} 行")
                self.tab_widget.currentWidget().load_data()

            except Exception as e:
                QMessageBox.critical(self, "错误", f"删除失败: {str(e)}")
            finally:
                if db is not None:
                    try:
                        db.close()
                    except:
                        # 忽略关闭连接时的错误
                        pass


    def search_in_post_list(self, task_ids: List[str]):
        """在帖子列表中搜索任务ID"""
        self._search_in_tab_by_name(task_ids, "帖子列表")
    
    def search_in_detail_list(self, task_ids: List[str]):
        """在帖子详情中搜索任务ID"""
        self._search_in_tab_by_name(task_ids, "帖子详情")
    
    def search_in_score_list(self, task_ids: List[str]):
        """在虎扑评分中搜索任务ID"""
        self._search_in_tab_by_name(task_ids, "虎扑评分")
    
    def _search_in_tab_by_name(self, task_ids: List[str], tab_name_keyword: str):
        """通用搜索函数：根据标签页名称关键字搜索任务ID"""
        try:
            if not task_ids:
                self._show_warning("请选择要搜索的任务")
                return
            
            # 获取父级DbTableViewer
            parent_viewer = self._get_parent_viewer()
            if parent_viewer is None:
                return
            
            # 查找目标标签页
            for i in range(parent_viewer.tab_widget.count()):
                tab_text = parent_viewer.tab_widget.tabText(i)
                if tab_name_keyword in tab_text:
                    parent_viewer.tab_widget.setCurrentIndex(i)
                    target_tab = parent_viewer.tab_widget.currentWidget()
                    
                    # 等待标签页加载完成
                    QTimer.singleShot(500, lambda: self._apply_task_id_filter(target_tab, task_ids))
                    return
            
            self._show_warning(f"未找到{tab_name_keyword}标签页")
        except Exception as e:
            self._show_critical_error("搜索失败", str(e))
    
    def _apply_task_id_filter(self, tab_widget, task_ids: List[str]):
        """应用任务ID筛选"""
        try:
            if hasattr(tab_widget, 'search_combo1') and hasattr(tab_widget, 'search_input1'):
                # 设置筛选条件为任务ID
                index = tab_widget.search_combo1.findText("任务ID")
                if index >= 0:
                    tab_widget.search_combo1.setCurrentIndex(index)
                    # 如果有多个任务ID，用逗号分隔
                    task_id_str = ", ".join(task_ids)
                    tab_widget.search_input1.setText(task_id_str)
                    tab_widget.search_input2.clear()
                    tab_widget.apply_search()
            else:
                self._show_warning("目标标签页不支持搜索功能")
        except Exception as e:
            self._show_critical_error("应用搜索失败", str(e))
    
    def _show_warning(self, message: str):
        """显示警告消息的通用函数"""
        QMessageBox.warning(self, "警告", message)
    
    def _show_critical_error(self, title: str, message: str):
        """显示错误消息的通用函数"""
        QMessageBox.critical(self, title, message)
    
    def _get_parent_viewer(self):
        """获取父级DbTableViewer的通用函数"""
        parent_viewer = self.get_parent_viewer()
        if not parent_viewer:
            self._show_warning("无法访问标签页控件")
            return None
        return parent_viewer
    
    def _set_fixed_widths(self, widgets_with_widths: Dict):
        """批量设置组件固定宽度的通用函数"""
        for widget, width in widgets_with_widths.items():
            if widget is not None:
                widget.setFixedWidth(width)
    
    def _create_button(self, text: str, width: int = None, tooltip: str = None):
        """创建按钮的通用函数"""
        btn = QPushButton(text)
        if width:
            btn.setFixedWidth(width)
        if tooltip:
            btn.setToolTip(tooltip)
        return btn
    
    def _create_combo_box(self, width: int = None, items: List[str] = None):
        """创建下拉框的通用函数"""
        combo = QComboBox()
        if width:
            combo.setFixedWidth(width)
        if items:
            combo.addItems(items)
        return combo
    
    def _create_line_edit(self, width: int = None, placeholder: str = None):
        """创建文本输入框的通用函数"""
        line_edit = QLineEdit()
        if width:
            line_edit.setFixedWidth(width)
        if placeholder:
            line_edit.setPlaceholderText(placeholder)
        return line_edit


# ------------------------------ 快捷创建标签页配置的工具函数 - 保持不变 ------------------------------
def create_tab_config(tab_name: str, table_name: str,
                      columns_to_display: List[str],
                      column_aliases: Optional[Dict] = None,
                      column_width_config: Optional[Dict] = None,
                      context_menu_actions: Optional[List[Dict]] = None,
                      db_path: Optional[str] = None) -> Dict:
    """
    快捷创建标签页配置的工具函数
    :param tab_name: 标签页显示名称
    :param table_name: 数据库表名
    :param columns_to_display: 要显示的列
    :param column_aliases: 列别名映射
    :param column_width_config: 列宽配置 {列名: (模式, 宽度)}
    :param context_menu_actions: 右键菜单配置
    :param db_path: 数据库路径（可选，如果不指定则使用默认数据库）
    :return: 完整的标签页配置
    """
    # 默认列别名
    default_aliases = {
        "id": "ID",
        "shop_name": "店铺名称",
        "shop_abbr": "缩写",
        "phone": "手机号",
        "password": "密码",
        "connect_status": "连接状态",
        "create_time": "创建时间",
        "update_time": "更新时间",
        "headers": "请求头",
        "cookies": "Cookies",
        "task_id": "任务ID",
        "func_name": "函数名",
        "status": "状态",
        "msg": "消息"
    }

    # 默认列宽配置
    default_widths = {
        "id": ("ResizeToContents", 0),
        "phone": ("ResizeToContents", 0),
        "password": ("ResizeToContents", 0),
        "connect_status": ("ResizeToContents", 0),
        "shop_name": ("Fixed", 150),
        "shop_abbr": ("Fixed", 50),
        "create_time": ("Fixed", 180),
        "update_time": ("Fixed", 180),
        "headers": ("Fixed", 200),
        "cookies": ("Fixed", 200),
        "task_id": ("Fixed", 250),
        "task_name": ("Fixed", 100),
        "status": ("Fixed", 60),
        "msg": ("Fixed", 200),
    }

    return {
        "tab_name": tab_name,
        "table_name": table_name,
        "columns_to_display": columns_to_display,
        "column_aliases": column_aliases or default_aliases,
        "column_width_config": column_width_config or default_widths,
        "context_menu_actions": context_menu_actions or [],
        "db_path": db_path
    }


if __name__ == "__main__":
    app = QApplication(sys.argv)

    # 数据库配置
    DB_PATH = "../配置文件_系统配置/db_config.json"

    # 创建主窗口实例
    viewer = DbTableViewer(DB_PATH, [
        # 第一个标签页：店铺表
        create_tab_config(
            tab_name="店铺管理",
            table_name="shops",
            columns_to_display=["id", "shop_name", "shop_abbr", "phone", "password",
                                "connect_status", "create_time", "update_time", "headers", "cookies"],
            column_width_config={
                "id": ("Fixed", 60),
                "shop_name": ("Fixed", 150),
                "shop_abbr": ("Fixed", 100),
                "phone": ("Fixed", 120),
                "password": ("Fixed", 120),
                "connect_status": ("Fixed", 100),
                "create_time": ("Fixed", 150),
                "update_time": ("Fixed", 150),
                "headers": ("Fixed", 200),
                "cookies": ("Fixed", 200)
            },
            context_menu_actions=[
                {"修改手机号": lambda ids: viewer.modify_field("shops", "phone", "手机号", ids)},
                {"修改密码": lambda ids: viewer.modify_field("shops", "password", "密码", ids)},
                {"清空认证": lambda ids: viewer.clear_auth("shops", ids)},
                {"删除选中行": lambda ids: viewer.delete_rows("shops", ids)}
            ]
        ),
        # 第二个标签页：任务表（示例）
        create_tab_config(
            tab_name="任务管理",
            table_name="task",
            columns_to_display=["id", "task_name", "status", "func_name", "task_group", "msg", "remarks", "task_id", "ip", "create_time", "update_time"],
            column_aliases={
                "id": "ID",
                "task_name": "任务名称",
                "status": "状态",
                "func_name": "函数名称",
                "task_group": "任务组",
                "msg": "信息",
                "remarks": "备注",
                "task_id": "任务ID",
                "ip": "代理IP",
                "create_time": "创建时间",
                "update_time": "更新时间"
            },
            column_width_config={
                "id": ("Fixed", 60),
                "task_name": ("Fixed", 200),
                "status": ("Fixed", 80),
                "func_name": ("Fixed", 200),
                "task_group": ("Fixed", 150),
                "msg": ("Fixed", 200),
                "remarks": ("Fixed", 200),
                "task_id": ("Fixed", 200),
                "ip": ("Fixed", 120),
                "create_time": ("Fixed", 150),
                "update_time": ("Fixed", 150)
            },
                    context_menu_actions=[
                        {"修改状态": lambda ids: viewer.modify_task_status(ids)},
                        {"删除任务": lambda ids: viewer.delete_task_rows(ids)}
                    ]
        )
    ])

    viewer.show()
    sys.exit(app.exec_())